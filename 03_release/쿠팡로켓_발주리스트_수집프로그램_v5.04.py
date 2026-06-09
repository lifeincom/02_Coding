import os
import pandas as pd
from openpyxl import load_workbook
from tksheet import Sheet
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Listbox, Frame, Label, Button, Entry, Scrollbar, LabelFrame, Toplevel
import warnings
import matplotlib.pyplot as plt

plt.rc('font', family='NanumGothic')  # For Windows

with warnings.catch_warnings(record=True):
    warnings.simplefilter("always")


# ===================== 상수 정의 =====================
class Constants:
    """프로그램에서 사용하는 상수 정의"""
    # 경로 설정
    DIR_PATHS = [
        r"\\NAS451\team451",
        r"\\192.168.0.101\team451",
        r"D:\hSync\Coding"
    ]
    SKU_LIST_FILE = r"DB\쿠팡SKU리스트.xlsx"
    ORDER_LIST_DIR = r"05-쿠팡로켓배송발주관리\발주(출고)리스트"
    
    # Excel 파일 구조 상수
    ORDER_NO_ROW = 10
    ORDER_NO_COL = 3
    DATE_ROW = 13
    DATE_COL = 6
    CARGO_COL = 3
    DATA_START_ROW = 22
    
    # Excel 셀 위치
    ITEM_NO_COL = 1
    SKU_ID_COL = 2
    SKU_NAME_COL = 3
    CENTER_COL = 6
    ORDER_QTY_COL = 7
    ORDER_PRICE_COL = 10
    BARCODE_ROW_OFFSET = 1
    BARCODE_COL = 3
    
    # 컬럼 리스트
    COL_LIST = [
        "발주번호", "정렬NO", "발주서NO", "박스NO", "SKU ID", "SKU 이름", 
        "옵션코드", "발주수량", "확정수량", "비고", "SKU Barcode", 
        "쿠팡옵션코드", "물류센터", "입고예정일", "출고예정일", 
        "발주공급가", "견적공급가"
    ]
    
    INFO_LIST = [
        "여자청바지", "사입리스트", "남자청바지", "티셔츠", "폴로바지", 
        "폴로티셔츠", "BHP", "아크시", "총발주수량"
    ]
    
    CATEGORY_ORDER = [
        ["여자청바지", "1"], ["사입리스트", "2"], ["남자청바지", "3"], 
        ["티셔츠", "4"], ["폴로바지", "5"], ["폴로티셔츠", "6"], 
        ["BHP", "7"], ["아크시", "8"], ["총발주수량", "9"]
    ]
    
    # GUI 설정
    WINDOW_TITLE = "쿠팡로켓 발주 수집 자동화 프로그램 v5.04"
    WINDOW_GEOMETRY = "1520x850+20+20"
    WINDOW_BG = "#F0F0F0"  # Slightly off-white for modern look
    BTN_RELIEF = "flat"
    BTN_FG = "#FFFFFF"
    BTN_BG = "#333333"
    BTN_BG_GRAY = "#555555"
    BTN_BG_BLUE = "#007BFF"  # Modern Blue
    BTN_BG_RED = "#DC3545"   # Modern Red
    
    # Sheet 설정
    SHEET_HEIGHT = 480
    SHEET_HEADER_HEIGHT = 25
    SHEET_HEADER_FG = "#FFFFFF"
    SHEET_HEADER_BG = "#111111"
    FONT_NAME = "NanumGothic"
    FONT_SIZE_HEADER = 10
    FONT_SIZE_NORMAL = 9


# ===================== 유틸리티 함수 =====================
def get_valid_path(paths):
    """유효한 경로를 찾아 반환"""
    for path in paths:
        if os.path.isdir(path):
            return path
    return None


# ===================== 데이터 처리 클래스 =====================
class OrderDataProcessor:
    """발주 데이터 처리 클래스"""
    
    def __init__(self, base_path):
        self.base_path = base_path
        self.current_df = pd.DataFrame(columns=Constants.COL_LIST)
    
    def extract_order_list(self, files):
        """발주서 파일에서 발주 리스트 추출"""
        order_df = pd.DataFrame(columns=Constants.COL_LIST)
        sort_idx = 1
        row_idx = 0
        
        for file in files:
            try:
                workbook = load_workbook(file)
                worksheet = workbook.active
                
                order_no = int(worksheet.cell(
                    row=Constants.ORDER_NO_ROW, 
                    column=Constants.ORDER_NO_COL
                ).value)
                
                date_value = worksheet.cell(
                    row=Constants.DATE_ROW, 
                    column=Constants.DATE_COL
                ).value
                delivery_date = str(date_value)[:10] if date_value else ""
                
                for row in range(Constants.DATA_START_ROW, worksheet.max_row + 1):
                    item_no = worksheet.cell(row=row, column=Constants.ITEM_NO_COL).value
                    
                    if item_no is not None:
                        sku_id = int(worksheet.cell(row=row, column=Constants.SKU_ID_COL).value)
                        sku_name = str(worksheet.cell(row=row, column=Constants.SKU_NAME_COL).value)
                        order_qty = int(worksheet.cell(row=row, column=Constants.ORDER_QTY_COL).value)
                        barcode = str(worksheet.cell(
                            row=row + Constants.BARCODE_ROW_OFFSET, 
                            column=Constants.BARCODE_COL
                        ).value)
                        center = str(worksheet.cell(row=row, column=Constants.CENTER_COL).value)
                        order_price = worksheet.cell(row=row, column=Constants.ORDER_PRICE_COL).value
                        
                        item_data = [
                            order_no, sort_idx, int(item_no), '', int(sku_id), sku_name, 
                            '', order_qty, '', '', barcode, '', center, delivery_date, 
                            '', order_price, ''
                        ]
                        order_df.loc[row_idx] = item_data
                        sort_idx += 1
                        row_idx += 1
                        
            except Exception as e:
                print(f"파일 처리 중 오류 발생 ({file}): {e}")
                continue
        
        return order_df
    
    def merge_with_sku(self, order_df):
        """SKU 리스트와 병합"""
        sku_list_path = os.path.join(self.base_path, Constants.SKU_LIST_FILE)
        
        if not os.path.exists(sku_list_path):
            raise FileNotFoundError(f"SKU 리스트 파일을 찾을 수 없습니다: {sku_list_path}")
        
        try:
            # 필요한 컬럼만 선택
            order_cols = [
                "발주번호", "정렬NO", "발주서NO", "박스NO", "SKU ID", "SKU 이름", 
                "발주수량", "확정수량", "비고", "SKU Barcode", "물류센터", 
                "입고예정일", "출고예정일", "발주공급가"
            ]
            order_subset = order_df.loc[:, order_cols]
            
            # SKU 리스트 읽기
            sku_df = pd.read_excel(sku_list_path, skiprows=1)
            sku_cols = ["SKU ID", "쿠팡옵션코드", "옵션코드", "견적공급가", "상품분류"]
            sku_subset = sku_df.loc[:, sku_cols]
            
            # 병합
            merged_df = order_subset.merge(sku_subset, on="SKU ID", how="left")
            result_df = merged_df.reindex(columns=Constants.COL_LIST)
            
            self.current_df = result_df
            return result_df, merged_df
            
        except Exception as e:
            print(f"SKU 병합 중 오류 발생: {e}")
            return order_df, order_df
    
    def get_category_summary(self, merged_df):
        """카테고리별 요약 정보 생성"""
        category_summary = merged_df.groupby("상품분류", as_index=False)["발주수량"].agg("sum")
        
        category_order_df = pd.DataFrame(
            Constants.CATEGORY_ORDER, 
            columns=["상품분류", "순서"]
        )
        
        result_df = pd.merge(
            category_order_df, 
            category_summary, 
            how="left"
        ).drop("순서", axis=1).fillna(0)
        
        # 총발주수량 설정
        result_df.iloc[8, 1] = merged_df["발주수량"].sum()
        
        # 전치하여 리스트로 변환
        transposed = result_df.T
        volumes = [int(transposed.iloc[1, x]) for x in range(9)]
        
        return volumes


# ===================== 로딩 화면 클래스 =====================
class LoadingWindow:
    """로딩 화면 클래스"""
    
    def __init__(self, parent, message="처리 중입니다..."):
        self.parent = parent
        self.window = Toplevel(parent)
        self.window.title("처리 중")
        self.window.geometry("400x150")
        self.window.resizable(False, False)
        self.window.configure(bg=Constants.WINDOW_BG)
        
        # 부모 창 중앙에 위치
        self.window.transient(parent)
        self.window.grab_set()
        
        # 부모 창 중앙 계산
        parent_x = parent.winfo_x()
        parent_y = parent.winfo_y()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        window_width = 400
        window_height = 150
        
        x = parent_x + (parent_width // 2) - (window_width // 2)
        y = parent_y + (parent_height // 2) - (window_height // 2)
        
        self.window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 메시지 레이블
        self.message_label = Label(
            self.window, 
            text=message, 
            font=(Constants.FONT_NAME, 12, 'normal'),
            bg=Constants.WINDOW_BG,
            fg="#333333"
        )
        self.message_label.pack(pady=30)
        
        # 진행 표시 (점 애니메이션)
        self.dots_label = Label(
            self.window,
            text=".",
            font=(Constants.FONT_NAME, 20, 'normal'),
            bg=Constants.WINDOW_BG,
            fg=Constants.BTN_BG_BLUE
        )
        self.dots_label.pack()
        
        self.dots_count = 0
        self.animate_dots()
        
        # 화면 업데이트
        self.window.update()
    
    def animate_dots(self):
        """점 애니메이션"""
        dots = "." * ((self.dots_count % 3) + 1)
        self.dots_label.config(text=dots)
        self.dots_count += 1
        self.window.after(500, self.animate_dots)
    
    def update_message(self, message):
        """메시지 업데이트"""
        self.message_label.config(text=message)
        self.window.update()
    
    def close(self):
        """로딩 창 닫기"""
        self.window.destroy()


# ===================== 파일명 변경 클래스 =====================
class FileNameChanger:
    """발주서 파일명 변경 클래스"""
    
    @staticmethod
    def generate_new_filename(file_path):
        """새로운 파일명 생성"""
        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            date_value = str(worksheet.cell(
                row=Constants.DATE_ROW, 
                column=Constants.DATE_COL
            ).value)
            date_str = date_value[5:7] + date_value[8:10]
            
            cargo = str(worksheet.cell(
                row=Constants.DATE_ROW, 
                column=Constants.CARGO_COL
            ).value)
            
            order_no = str(worksheet.cell(
                row=Constants.ORDER_NO_ROW, 
                column=Constants.ORDER_NO_COL
            ).value)
            
            return f"{date_str}_{cargo}_{order_no}.xlsx"
            
        except Exception as e:
            raise ValueError(f"파일명 생성 중 오류: {e}")
    
    @staticmethod
    def rename_file(source_path, dest_path):
        """파일명 변경"""
        if os.path.exists(dest_path):
            raise FileExistsError(f"동일한 파일명이 이미 존재합니다: {dest_path}")
        os.rename(source_path, dest_path)


# ===================== GUI 클래스 =====================
class OrderListApp:
    """발주 리스트 정리 애플리케이션"""
    
    def __init__(self, root, base_path):
        self.root = root
        self.base_path = base_path
        self.processor = OrderDataProcessor(base_path)
        self.setup_window()
        self.create_widgets()
        self.show_frame(self.frm1)
    
    def setup_window(self):
        """윈도우 기본 설정"""
        self.root.title(Constants.WINDOW_TITLE)
        self.root.geometry(Constants.WINDOW_GEOMETRY)
        self.root.state('zoomed')
        self.root.minsize(1520, 850)
        self.root.resizable(True, True)
    
    def create_widgets(self):
        """위젯 생성"""
        self.create_header()
        self.create_container()
        self.create_frame1()
        self.create_frame2()
        self.create_bottom_bar()
    
    def create_bottom_bar(self):
        """하단 상태 메시지 바 생성"""
        self.frm_bottom = Frame(self.root, bg=Constants.SHEET_HEADER_BG, height=30)
        self.frm_bottom.pack(side="bottom", fill="x")
        
        self.status_label = Label(
            self.frm_bottom, 
            text="준비", 
            font=(Constants.FONT_NAME, 10, 'normal'),
            bg=Constants.SHEET_HEADER_BG,
            fg="#FFFFFF",
            anchor="w"
        )
        self.status_label.pack(fill="both", padx=10, pady=5)

    def update_status(self, message, is_error=False):
        """상태 메시지 업데이트"""
        color = "#FF4000" if is_error else "#FFFFFF"
        self.status_label.config(text=message, fg=color)
        self.root.update()

    def create_header(self):
        """헤더 프레임 생성"""
        self.frm_header = Frame(self.root, bg=Constants.SHEET_HEADER_BG, height=50)
        self.frm_header.pack(fill="x")
        
        # Header Content Frame to center vertically if needed or padding
        header_content = Frame(self.frm_header, bg=Constants.SHEET_HEADER_BG)
        header_content.pack(fill="both", expand=True, padx=20, pady=10)

        header_text = (
            "★ 프로그램 정보 : 상품정보 수집 오류 시 연동데이터 파일를 확인해 주세요 / "
            "프로그램 연동데이터 : 서버 NAS451/team451/DB/쿠팡SKU리스트.xlsx"
        )
        label_header = Label(
            header_content, 
            text=header_text, 
            bg=Constants.SHEET_HEADER_BG, 
            fg="#FFFFFF", 
            font=(Constants.FONT_NAME, 9),
            anchor="w"
        )
        label_header.pack(side="left", fill="x")
        
        # Button Frame (Right side)
        btn_frame = Frame(header_content, bg=Constants.SHEET_HEADER_BG)
        btn_frame.pack(side="right")

        btn_close = Button(
            btn_frame, text="프로그램 종료", width=15, cursor="hand2",
            relief=Constants.BTN_RELIEF, fg=Constants.BTN_FG, bg=Constants.BTN_BG, 
            font=(Constants.FONT_NAME, 9, "bold"),
            command=self.root.quit
        )
        btn_close.pack(side="right", padx=5)
        
        btn_frm1 = Button(
            btn_frame, text="발주서 리스트 정리", width=18, cursor="hand2",
            relief=Constants.BTN_RELIEF, fg=Constants.BTN_FG, bg=Constants.BTN_BG_BLUE,
            font=(Constants.FONT_NAME, 9, "bold"), 
            command=lambda: self.show_frame(self.frm1)
        )
        btn_frm1.pack(side="right", padx=5)
        
        btn_frm2 = Button(
            btn_frame, text="발주서 파일명 일괄 변경", width=22, cursor="hand2",
            relief=Constants.BTN_RELIEF, fg=Constants.BTN_FG, bg=Constants.BTN_BG_RED, 
            font=(Constants.FONT_NAME, 9, "bold"),
            command=lambda: self.show_frame(self.frm2)
        )
        btn_frm2.pack(side="right", padx=5)
    
    def create_container(self):
        """컨테이너 프레임 생성"""
        self.frm_container = Frame(self.root)
        self.frm_container.pack(fill="both", padx=10, pady=5, ipady=5, expand=True)
        
        self.frm1 = Frame(self.frm_container)
        self.frm2 = Frame(self.frm_container)
        
        for frame in (self.frm1, self.frm2):
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
    
    def create_frame1(self):
        """프레임1 (발주서 리스트 정리) 생성"""
        # ================= Top Area (File List + Controls) =================
        frm1_top = Frame(self.frm1, bg=Constants.WINDOW_BG)
        frm1_top.pack(fill="x", padx=20, pady=15)
        
        # --- Left: File List ---
        frm1_left = Frame(frm1_top, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm1_left.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Header for List
        # Using a Frame to simulate a header bar if needed, or just padding
        # But the image shows a box. I'll just use the listbox inside.
        
        frm1_scrollbar = Scrollbar(frm1_left)
        frm1_scrollbar.pack(side="right", fill="y", padx=2, pady=2)
        
        self.frm1_load_file_list_box = Listbox(
            frm1_left, 
            selectmode="extended", 
            height=8, 
            yscrollcommand=frm1_scrollbar.set,
            relief="flat",
            width=100,
            bd=0
        )
        self.frm1_load_file_list_box.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        frm1_scrollbar.config(command=self.frm1_load_file_list_box.yview)
        
        # --- Right: Controls & Info ---
        frm1_right = Frame(frm1_top, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm1_right.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        # 1. Action Buttons Row
        frm1_btns = Frame(frm1_right, bg=Constants.WINDOW_BG)
        frm1_btns.pack(fill="x", padx=10, pady=15)
        
        btn_style = {
            "relief": Constants.BTN_RELIEF,
            "bg": Constants.BTN_BG,
            "fg": Constants.BTN_FG,
            "font": (Constants.FONT_NAME, 9),
            "cursor": "hand2"
        }
        
        # Buttons (Order: Init, Del, Copy, Run, Add) - Reversed for Right Packing
        # Or logical order L->R: Add -> Run -> Copy -> Del -> Init
        
        # Based on image: 1. Select, 2. Run, 3. Copy, Del, Init
        Button(frm1_btns, text="초기화", width=8, command=self.frm1_reset, **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="선택 삭제", width=10, command=self.frm1_del_file, **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="3. 발주서 수집 상품리스트 복사", width=25, command=self.frm1_order_list_copy, **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="2. 발주서 상품정보 수집 실행", width=25, command=self.frm1_start, **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="1. 발주서 파일 선택", width=18, command=self.frm1_add_file, **btn_style).pack(side="right", padx=5, ipady=3)
        
        # 2. Info Grid (Table)
        frm1_info = Frame(frm1_right, bg="#E0E0E0") # Grid line color
        frm1_info.pack(fill="x", padx=10, pady=(0, 15))
        
        # Info Grid Headers & Values placeholders
        # We need to store label references to update them later
        self.info_labels = {}
        
        cols = Constants.INFO_LIST
        
        # Configure grid columns
        for i in range(len(cols)):
            frm1_info.grid_columnconfigure(i, weight=1)
            
        # Headers
        for i, col in enumerate(cols):
            lbl = Label(
                frm1_info, text=col, 
                bg="#EEEEEE", fg="#333333", 
                font=(Constants.FONT_NAME, 9, "bold"),
                relief="flat", height=2, width=5
            )
            lbl.grid(row=0, column=i, sticky="nswe", padx=1, pady=1)
            
        # Values
        for i, col in enumerate(cols):
            lbl = Label(
                frm1_info, text="0", 
                bg="#FFFFFF", fg="#333333", 
                font=(Constants.FONT_NAME, 9),
                relief="flat", height=2, width=5
            )
            lbl.grid(row=1, column=i, sticky="nswe", padx=1, pady=(0, 1))
            self.info_labels[col] = lbl

        # ================= Bottom Area (Sheet) =================
        frm1_contents = Frame(self.frm1, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm1_contents.pack(fill="both", padx=20, pady=(0, 20), expand=True)
        
        self.frm1_xsheet = Frame(frm1_contents)
        self.frm1_xsheet.pack(fill="both", padx=5, pady=5, expand=True)
        
        # 초기 빈 시트 표시
        self.output_sheet(self.frm1_xsheet, pd.DataFrame(columns=Constants.COL_LIST))
    
    def create_frame2(self):
        """프레임2 (파일명 변경) 생성"""
        # ================= Top Control Bar =================
        frm2_top = Frame(self.frm2, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_top.pack(fill="x", padx=20, pady=15)
        
        # Title Label
        desc_text = (
            "★ 파일명 변경 : 쿠팡 로켓배송 발주서 파일을 선택해 주세요. "
            "(변경 파일명 형식 : 입고예정일_물류센터_발주서번호)"
        )
        frm2_desc_label = Label(
            frm2_top, 
            text=desc_text, 
            font=(Constants.FONT_NAME, 9),
            bg=Constants.WINDOW_BG,
            anchor="w"
        )
        frm2_desc_label.pack(side="left", padx=10, pady=10)
        
        # Buttons (Right side of Top Bar)
        
        # Common button style
        btn_style = {
            "relief": Constants.BTN_RELIEF,
            "bg": Constants.BTN_BG,
            "fg": Constants.BTN_FG,
            "font": (Constants.FONT_NAME, 9),
            "cursor": "hand2"
        }

        # 초기화
        btn_reset = Button(frm2_top, text="초기화", width=8, command=self.frm2_reset, **btn_style)
        btn_reset.pack(side="right", padx=5, pady=8, ipady=3)

        # 4. 선택 삭제
        btn_del = Button(frm2_top, text="선택 삭제", width=12, command=self.frm2_del_file, **btn_style)
        btn_del.pack(side="right", padx=5, pady=8, ipady=3)

        # 3. 파일명 변경 실행
        btn_run = Button(frm2_top, text="3. 파일명 변경 실행", width=18, command=self.frm2_change_filename, **btn_style)
        btn_run.pack(side="right", padx=5, pady=8, ipady=3)

        # 2. 저장 폴더 선택
        btn_dir = Button(frm2_top, text="2. 저장 폴더 선택", width=15, command=self.frm2_output_dir_path, **btn_style)
        btn_dir.pack(side="right", padx=5, pady=8, ipady=3)

        # 1. 변경파일 선택
        btn_add = Button(frm2_top, text="1. 변경파일 선택", width=15, command=self.frm2_add_file, **btn_style)
        btn_add.pack(side="right", padx=5, pady=8, ipady=3)

        # ================= Path Bar =================
        frm2_path_bar = Frame(self.frm2, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_path_bar.pack(fill="x", padx=20, pady=(0, 15))
        
        lbl_path = Label(
            frm2_path_bar, 
            text="변경파일 저장 폴더", 
            font=(Constants.FONT_NAME, 9, "bold"), 
            bg=Constants.WINDOW_BG
        )
        lbl_path.pack(side="left", padx=15, pady=10)
        
        self.frm2_txt_save_path = Entry(frm2_path_bar, relief="flat", bg="#F9F9F9")
        self.frm2_txt_save_path.pack(side="left", fill="x", expand=True, padx=(0, 15), pady=10, ipady=3)

        # ================= Content Area (Split Lists) =================
        frm2_contents = Frame(self.frm2, bg=Constants.WINDOW_BG)
        frm2_contents.pack(fill="both", padx=20, pady=(0, 20), expand=True)

        # Left List (Original)
        frm2_sector1 = Frame(frm2_contents, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_sector1.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Right List (Changed)
        frm2_sector2 = Frame(frm2_contents, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_sector2.pack(side="right", fill="both", expand=True, padx=(10, 0))

        # Listboxes with Scrollbars
        def create_listbox(parent):
            sb = Scrollbar(parent)
            sb.pack(side="right", fill="y", padx=2, pady=2)
            lb = Listbox(
                parent, 
                selectmode="extended", 
                yscrollcommand=sb.set,
                relief="flat",
                bd=0,
                highlightthickness=0
            )
            lb.pack(side="left", fill="both", expand=True, padx=5, pady=5)
            sb.config(command=lb.yview)
            return lb

        self.frm2_original_file = create_listbox(frm2_sector1)
        self.frm2_changed_file = create_listbox(frm2_sector2)
    
    def show_frame(self, frame):
        """프레임 전환"""
        frame.tkraise()
    
    def output_sheet(self, frame, dataframe):
        """시트 출력"""
        for widget in frame.winfo_children():
            widget.destroy()
        
        sheet_data = dataframe.values.tolist()
        col_list = list(dataframe.columns)
        
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)
        
        inner_frame = Frame(frame)
        inner_frame.grid_columnconfigure(0, weight=1)
        inner_frame.grid_rowconfigure(0, weight=1)
        
        sheet_widget = Sheet(
            inner_frame, data=sheet_data, height=Constants.SHEET_HEIGHT, 
            headers=col_list, header_height=Constants.SHEET_HEADER_HEIGHT, 
            header_fg=Constants.SHEET_HEADER_FG, header_bg=Constants.SHEET_HEADER_BG
        )
        sheet_widget.header_font((Constants.FONT_NAME, Constants.FONT_SIZE_HEADER, 'normal'))
        sheet_widget.font((Constants.FONT_NAME, Constants.FONT_SIZE_NORMAL, 'normal'))
        sheet_widget.table_align(align="w")
        sheet_widget["A:E"].align("c")
        sheet_widget["H:K"].align("c")
        sheet_widget["M:O"].align("c")
        sheet_widget["P:Q"].align("e")
        sheet_widget.set_all_column_widths(
            width=None, only_set_if_too_small=False, 
            redraw=True, recreate_selection_boxes=True
        )
        sheet_widget.enable_bindings()
        
        inner_frame.grid(row=0, column=0, sticky="nswe")
        sheet_widget.grid(row=0, column=0, sticky="nswe")
        inner_frame.pack(fill="both", expand=True)
    
    # ===================== Frame1 이벤트 핸들러 =====================
    def frm1_add_file(self):
        """파일 추가"""
        initial_dir = os.path.join(
            self.base_path, Constants.ORDER_LIST_DIR
        ) if self.base_path else ""
        
        file_list = filedialog.askopenfilenames(
            title="발주서를 선택해 주세요.",
            filetypes=(("xlsx 파일", "*.xlsx"), ("모든 파일", "*.*")),
            initialdir=initial_dir
        )
        
        for file_path in file_list:
            self.frm1_load_file_list_box.insert(tk.END, file_path)
    
    def frm1_reset(self):
        """초기화"""
        self.frm1_load_file_list_box.delete(0, tk.END)
        # Assuming self.info_labels is initialized in create_frame1
        if hasattr(self, 'info_labels'):
            for lbl in self.info_labels.values():
                lbl.config(text="0")
        self.output_sheet(self.frm1_xsheet, pd.DataFrame(columns=Constants.COL_LIST))
    
    def frm1_del_file(self):
        """선택 파일 삭제"""
        for index in reversed(self.frm1_load_file_list_box.curselection()):
            self.frm1_load_file_list_box.delete(index)
    
    def frm1_start(self):
        """발주서 수집 시작"""
        if self.frm1_load_file_list_box.size() == 0:
            messagebox.showwarning("경고", "발주서 파일을 추가하세요.")
            return
        
        self.frm1_order_list_ext()
    
    def frm1_order_list_ext(self):
        """발주 리스트 추출 및 처리"""
        files = list(self.frm1_load_file_list_box.get(0, tk.END))
        
        # 로딩 화면 표시
        loading = LoadingWindow(self.root, "발주서 파일을 읽는 중입니다...")
        
        try:
            loading.update_message("발주서 파일을 읽는 중입니다...")
            self.update_status("발주서 파일을 읽는 중입니다...")
            self.root.update()
            
            # 발주 리스트 추출
            order_df = self.processor.extract_order_list(files)
            
            loading.update_message("SKU 리스트와 병합 중입니다...")
            self.update_status("SKU 리스트와 병합 중입니다...")
            self.root.update()
            
            # SKU 병합
            result_df, merged_df = self.processor.merge_with_sku(order_df)
            
            loading.update_message("데이터를 정리하는 중입니다...")
            self.update_status("데이터를 정리하는 중입니다...")
            self.root.update()
            
            # 시트 출력
            self.output_sheet(self.frm1_xsheet, result_df)
            
            # 카테고리 요약
            volumes = self.processor.get_category_summary(merged_df)
            
            # Update Info Grid Labels
            if hasattr(self, 'info_labels'):
                for col, vol in zip(Constants.INFO_LIST, volumes):
                    if col in self.info_labels:
                        self.info_labels[col].config(text=str(vol))
            
            # 로딩 화면 닫기
            loading.close()
            
            self.update_status("발주리스트 수집이 완료 되었습니다.")
            messagebox.showinfo("알림", "발주리스트 수집이 완료 되었습니다.")
            
        except Exception as e:
            loading.close()
            self.update_status(f"오류 발생: {e}", is_error=True)
            messagebox.showerror("오류", f"발주리스트 수집 중 오류 발생: {e}")
    
    def frm1_order_list_copy(self):
        """발주 리스트 클립보드 복사"""
        if self.frm1_load_file_list_box.size() == 0:
            messagebox.showwarning("경고", "발주서 상품정보가 수집되지 않았습니다.")
            return
        
        try:
            if self.processor.current_df.empty:
                messagebox.showwarning("경고", "복사할 데이터가 없습니다.")
                return
            
            self.processor.current_df.to_clipboard(index=False, header=False)
            messagebox.showinfo("알림", "발주리스트 복사가 완료되었습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"복사 중 오류 발생: {e}")
    
    # ===================== Frame2 이벤트 핸들러 =====================
    def frm2_add_file(self):
        """파일 추가"""
        file_list = filedialog.askopenfilenames(
            title="변경 할 파일을 선택하세요",
            filetypes=(("XLSX 파일", "*.xlsx"), ("XLS 파일", "*.xls"), ("모든 파일", "*.*")),
            initialdir="다운로드"
        )
        
        for file_path in file_list:
            self.frm2_original_file.insert(tk.END, file_path)
    
    def frm2_reset(self):
        """초기화"""
        self.frm2_original_file.delete(0, tk.END)
        self.frm2_changed_file.delete(0, tk.END)
        self.frm2_txt_save_path.delete(0, tk.END)
    
    def frm2_del_file(self):
        """선택 파일 삭제"""
        for index in reversed(self.frm2_original_file.curselection()):
            self.frm2_original_file.delete(index)
    
    def frm2_output_dir_path(self):
        """저장 경로 선택"""
        output_path = filedialog.askdirectory(title="저장 할 폴더를 선택해 주세요")
        if output_path:
            self.frm2_txt_save_path.delete(0, tk.END)
            self.frm2_txt_save_path.insert(0, output_path)
    
    def frm2_change_filename(self):
        """파일명 변경 실행"""
        if self.frm2_original_file.size() == 0:
            messagebox.showwarning("경고", "파일명을 변경 할 파일을 추가하세요.")
            return
        
        if not self.frm2_txt_save_path.get():
            messagebox.showwarning("경고", "저장 경로를 선택하세요")
            return
        
        files = list(self.frm2_original_file.get(0, tk.END))
        output_path = self.frm2_txt_save_path.get()
        
        # 로딩 화면 표시
        loading = LoadingWindow(self.root, "파일명을 변경하는 중입니다...")
        
        try:
            processed_count = 0
            total_files = len(files)
            
            for file_path in files:
                try:
                    processed_count += 1
                    msg = f"파일명 변경 중... ({processed_count}/{total_files})"
                    loading.update_message(msg)
                    self.update_status(msg)
                    self.root.update()
                    
                    new_filename = FileNameChanger.generate_new_filename(file_path)
                    dest_path = os.path.join(output_path, new_filename)
                    
                    FileNameChanger.rename_file(file_path, dest_path)
                    
                    self.frm2_changed_file.insert(tk.END, dest_path)
                    
                except FileExistsError as e:
                    loading.close()
                    messagebox.showwarning("경고", str(e))
                    return
                except Exception as e:
                    loading.close()
                    messagebox.showerror("오류", f"파일 처리 중 오류 발생 ({file_path}): {e}")
                    return
            
            # 로딩 화면 닫기
            loading.close()
            self.update_status("파일명 변경이 완료 되었습니다.")
            messagebox.showinfo('알림', "파일명 변경이 완료 되었습니다.")
            
        except Exception as e:
            loading.close()
            self.update_status(f"파일명 변경 중 오류 발생: {e}", is_error=True)
            messagebox.showerror("오류", f"파일명 변경 중 오류 발생: {e}")


# ===================== 메인 실행 =====================
def main():
    """메인 함수"""
    base_path = get_valid_path(Constants.DIR_PATHS)
    
    if base_path is None:
        print("유효한 경로를 찾을 수 없습니다.")
        base_path = ""
    
    root = tk.Tk()
    app = OrderListApp(root, base_path)
    root.mainloop()


if __name__ == "__main__":
    main()
