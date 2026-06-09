"""
쿠팡 로켓배송 발주 수집·박스라벨 출력 프로그램 v6.07
=======================================================
[프로그램 개요]
  - 쿠팡 로켓배송 발주서(Excel)를 선택해 발주 리스트를 자동 수집하고,
    SKU 마스터 데이터와 병합한 뒤 결과를 GUI 시트로 표시한다.
  - 발주서 파일명을 '입고예정일_물류센터_발주번호' 형식으로 일괄 변경한다.
  - 박스 분류가 완료된 발주 리스트를 읽어 박스라벨 출력용 Excel을 생성한다.

[주요 라이브러리]
  - tkinter      : GUI 프레임워크 (창, 버튼, 리스트박스 등)
  - tksheet      : 스프레드시트 형태의 데이터 뷰어 위젯
  - pandas       : 데이터프레임 기반 데이터 처리
  - openpyxl     : Excel 파일 읽기·쓰기
  - matplotlib   : 한글 폰트 설정용 (차트 렌더링은 미사용)

[클래스 구조]
  Constants           → 프로그램 전체에서 공유하는 상수 모음
  OrderDataProcessor  → 발주 데이터 추출·병합·요약 로직
  LoadingWindow       → 비동기 처리 중 표시되는 로딩 팝업
  FileNameChanger     → 발주서 파일명 생성 및 변경 유틸
  OrderListApp        → 메인 GUI 애플리케이션 (tkinter 기반 3-Frame 구조)
"""

# ─────────────────── 표준 라이브러리 임포트 ───────────────────
import math           # 파렛트 수량 계산 시 올림(math.ceil) 사용
import os             # 경로 존재 확인, 파일 이름 변경
import re             # 파일명에서 특수문자 제거용 정규표현식
import warnings       # openpyxl 경고 억제

from copy import copy          # openpyxl 셀 스타일 복사 시 얕은 복사 사용
from pathlib import Path       # 플랫폼 독립적 경로 처리

# ─────────────────── 서드파티 라이브러리 임포트 ───────────────────
import matplotlib.pyplot as plt  # 한글 폰트 설정 목적으로 임포트
import pandas as pd              # 데이터프레임 처리

import tkinter as tk  # tkinter 루트 네임스페이스 (tk.END 등 상수 사용)
from openpyxl import Workbook, load_workbook       # Excel 파일 생성 및 로드
from openpyxl.worksheet.cell_range import CellRange  # 인쇄 영역 설정용
from tksheet import Sheet  # 스프레드시트 위젯 (pip install tksheet)
from tkinter import (
    Button, Entry, Frame, Label, Listbox, Scrollbar, Toplevel,
    filedialog, messagebox,  # 파일 대화상자, 메시지 팝업
)

# ─────────────────── 전역 설정 ───────────────────
# matplotlib 한글 폰트 지정 (Windows 환경: NanumGothic 사용)
# 리눅스/맥에서는 'AppleGothic' 또는 'Malgun Gothic' 등으로 변경 필요
plt.rc('font', family='NanumGothic')

# openpyxl이 발생시키는 스타일 관련 UserWarning 을 조용히 무시
with warnings.catch_warnings(record=True):
    warnings.simplefilter("always")


# ===================== 상수 정의 =====================
class Constants:
    """
    프로그램 전체에서 공유되는 상수 모음.

    매직 넘버나 하드코딩 문자열을 한 곳에서 관리하여
    유지보수 편의성을 높인다.

    [최적화 포인트]
      - 셀 위치 상수(ORDER_NO_ROW 등)를 이름 있는 상수로 분리해
        Excel 양식 변경 시 Constants 만 수정하면 전체 반영된다.
    """

    # ── 네트워크/로컬 경로 후보 목록 (위에서부터 순서대로 존재 여부 확인) ──
    DIR_PATHS = [
        r"\\NAS451\team451",         # 사무실 NAS (우선순위 1)
        r"\\192.168.0.101\team451",  # NAS IP 직접 접근 (우선순위 2)
        r"D:\hSync\03_Coding",          # 개발자 로컬 경로 (우선순위 3)
    ]

    # SKU 마스터 파일 상대경로 (DIR_PATHS 기준)
    SKU_LIST_FILE = r"DB\쿠팡SKU리스트.xlsx"

    # 발주리스트 저장 폴더 상대경로 (파일 탐색기 초기 위치로 사용)
    ORDER_LIST_DIR = r"05-쿠팡로켓배송발주관리\발주(출고)리스트"

    # 박스라벨 출력양식(템플릿) Excel 상대경로
    BOX_TEMPLATE_REL = r"05-쿠팡로켓배송발주관리\발주(출고)리스트\박스라벨_출력양식.xlsx"

    # ── 박스라벨 출력 관련 상수 ──
    # 템플릿 박스라벨_출력양식.xlsx 블록 구조 (15행 1블록):
    #   행1      : 여백 (height=9.95)
    #   행2      : 박스 적재리스트 제목 + 박스번호(D열)
    #   행3      : 입고예정일자 + 날짜/물류센터 정보(C열)
    #   행4      : 업 체 명 + 업체명(C열)
    #   행5      : 발주번호 + 발주번호값(C열)
    #   행6      : 컬럼 헤더 (NO / SKU No. / SKU NAME / 수량)
    #   행7~14   : 데이터 슬롯 (8행)
    #   행15     : 여백 (height=9.95)
    BOXES_PER_PALLET  = 16   # 파렛트 1개당 박스 최대 수량 (파렛트 수 계산에 사용)
    BOX_BLOCK_ROWS    = 15   # 템플릿 1개 박스 블록의 기본 행 수 (데이터 초과 시 동적 확장)
    BOX_HEADER_ROWS   = 6    # 블록 내 헤더 영역 행 수 (여백1+제목1+메타3+컬럼헤더1)
    BOX_FOOTER_ROWS   = 1    # 블록 내 하단 여백 행 수 (마지막 여백 행)
    # 기본 데이터 슬롯 = BOX_BLOCK_ROWS - BOX_HEADER_ROWS - BOX_FOOTER_ROWS = 8행
    BOX_SKU_NAME_MAX_LEN = 80  # SKU 이름 최대 길이 (초과 시 잘라냄)

    # 블록 내 헤더 셀 행 오프셋 (1-based, 블록 시작 행 기준)
    # 예) 블록 시작이 offset_row=1 이면:
    #   ROW_TITLE   = offset_row + 1  → 행2  (박스 적재리스트 + 박스번호)
    #   ROW_DATE    = offset_row + 2  → 행3  (입고예정일자 / 물류센터)
    #   ROW_COMPANY = offset_row + 3  → 행4  (업 체 명)
    #   ROW_ORDER   = offset_row + 4  → 행5  (발주번호)
    BOX_ROW_OFFSET_TITLE   = 1   # 박스 제목행 (박스번호 기입)
    BOX_ROW_OFFSET_DATE    = 2   # 날짜/물류센터 기입행
    BOX_ROW_OFFSET_COMPANY = 3   # 업체명 기입행 (고정값, 수정 불필요)
    BOX_ROW_OFFSET_ORDER   = 4   # 발주번호 기입행
    BOX_COL_BOX_NO         = 4   # 박스번호가 들어가는 열 (D열)
    BOX_COL_DATE_VALUE     = 3   # 날짜/물류센터 값이 들어가는 열 (C열)
    BOX_COL_ORDER_VALUE    = 3   # 발주번호 값이 들어가는 열 (C열)

    # 박스라벨 시트 컬럼별 픽셀 너비
    BOX_LABEL_COL_WIDTHS = {
        '발주번호':   110,
        '정렬NO':      65,
        '박스NO':      65,
        'SKU_ID':      95,
        'SKU_NAME':   720,
        '수량':        65,
        '물류센터':    85,
        '입고예정일': 105,
    }

    # 박스라벨 결과 요약 영역에 표시할 항목명
    BOX_RESULT_TITLES = ("박스수량", "파렛트수량", "출고상품수량")

    # ── 발주서 Excel 셀 위치 상수 (openpyxl 1-based 인덱스) ──
    ORDER_NO_ROW  = 10   # 발주번호가 있는 행
    ORDER_NO_COL  = 3    # 발주번호가 있는 열
    DATE_ROW      = 13   # 입고예정일이 있는 행
    DATE_COL      = 6    # 입고예정일이 있는 열
    CARGO_COL     = 3    # 물류센터(화물처) 값이 있는 열
    DATA_START_ROW = 22  # 품목 데이터가 시작되는 행

    # 품목 데이터 각 열 번호
    ITEM_NO_COL      = 1   # 순번
    SKU_ID_COL       = 2   # SKU ID
    SKU_NAME_COL     = 3   # SKU 이름
    CENTER_COL       = 6   # 물류센터
    ORDER_QTY_COL    = 7   # 발주수량
    ORDER_PRICE_COL  = 10  # 발주공급가
    BARCODE_ROW_OFFSET = 1  # 바코드 행은 품목 행 +1 에 위치
    BARCODE_COL      = 3   # 바코드 열

    # ── 발주 리스트 데이터프레임 컬럼 정의 (순서 고정) ──
    COL_LIST = [
        "발주번호", "정렬NO", "발주서NO", "박스NO", "SKU ID", "SKU 이름",
        "옵션코드", "발주수량", "확정수량", "비고", "SKU Barcode",
        "쿠팡옵션코드", "물류센터", "입고예정일", "출고예정일",
        "발주공급가", "견적공급가",
    ]

    # ── 카테고리 요약 정보 항목 이름 (발주서 리스트 정리 탭 상단 그리드) ──
    INFO_LIST = [
        "여자청바지", "사입리스트", "남자청바지", "티셔츠", "폴로바지",
        "폴로티셔츠", "BHP", "아크시", "총발주수량",
    ]

    # 카테고리와 정렬 순서 매핑 (groupby 후 지정 순서로 재정렬)
    CATEGORY_ORDER = [
        ["여자청바지", "1"], ["사입리스트", "2"], ["남자청바지", "3"],
        ["티셔츠", "4"],     ["폴로바지",   "5"], ["폴로티셔츠", "6"],
        ["BHP",     "7"],    ["아크시",     "8"], ["총발주수량", "9"],
    ]

    # ── GUI 스타일 상수 ──
    WINDOW_TITLE    = "쿠팡로켓 발주 수집·박스라벨 출력 프로그램 v6.07"
    WINDOW_GEOMETRY = "1520x850+20+20"  # 너비x높이+X위치+Y위치
    WINDOW_BG       = "#F0F0F0"  # 창 배경색 (밝은 회색)

    BTN_RELIEF  = "flat"     # 버튼 테두리 스타일 (납작한 평면형)
    BTN_FG      = "#FFFFFF"  # 버튼 글자색 (흰색)
    BTN_BG      = "#333333"  # 기본 버튼 배경색 (짙은 회색)
    BTN_BG_GRAY = "#555555"  # 보조 버튼 배경색
    BTN_BG_BLUE = "#007BFF"  # 주요 액션 버튼 (파란색)
    BTN_BG_RED  = "#DC3545"  # 경고성 액션 버튼 (빨간색)
    BTN_BG_GREEN = "#02C008" # 완료/출력 버튼 (초록색)

    # tksheet 스프레드시트 위젯 스타일
    SHEET_HEIGHT        = 480   # 시트 위젯 기본 높이 (픽셀)
    SHEET_HEADER_HEIGHT = 25    # 헤더 행 높이 (픽셀)
    SHEET_HEADER_FG     = "#FFFFFF"  # 헤더 글자색
    SHEET_HEADER_BG     = "#111111"  # 헤더 배경색 (거의 검정)

    FONT_NAME        = "NanumGothic"  # 프로그램 전체 사용 폰트
    FONT_SIZE_HEADER = 10             # 헤더 폰트 크기
    FONT_SIZE_NORMAL = 9              # 일반 텍스트 폰트 크기


# ===================== 유틸리티 함수 =====================

def get_valid_path(paths: list[str]) -> str | None:
    """
    후보 경로 목록 중 실제로 존재하는 첫 번째 디렉터리를 반환한다.

    Args:
        paths: 검사할 경로 문자열 목록 (Constants.DIR_PATHS)

    Returns:
        유효한 경로 문자열, 없으면 None

    사용 예:
        base = get_valid_path(Constants.DIR_PATHS)
        # → "\\\\NAS451\\team451"  (NAS가 마운트된 경우)
    """
    for path in paths:
        if os.path.isdir(path):
            return path
    return None


def get_box_template_path() -> str | None:
    """
    박스라벨 출력양식 Excel 파일의 절대 경로를 반환한다.

    Constants.DIR_PATHS 목록의 각 경로 아래에
    BOX_TEMPLATE_REL 상대경로를 결합해 파일 존재 여부를 순서대로 확인한다.

    Returns:
        파일이 존재하는 절대 경로 문자열, 없으면 None
    """
    for base in Constants.DIR_PATHS:
        path = Path(base) / Constants.BOX_TEMPLATE_REL
        if path.is_file():
            return str(path)
    return None


# ===================== 데이터 처리 클래스 =====================

class OrderDataProcessor:
    """
    발주서 Excel 파일에서 데이터를 추출하고, SKU 마스터와 병합하여
    최종 발주 데이터프레임을 생성하는 비즈니스 로직 클래스.

    GUI 와 완전히 분리되어 있으므로, 단독 테스트나 CLI 활용이 가능하다.

    Attributes:
        base_path (str)     : NAS 또는 로컬 루트 경로
        current_df (DataFrame): 가장 최근에 처리된 발주 데이터프레임
                                (클립보드 복사 기능에서 참조)
    """

    def __init__(self, base_path: str):
        self.base_path  = base_path
        # 초기값: 빈 데이터프레임 (컬럼 구조만 정의)
        self.current_df = pd.DataFrame(columns=Constants.COL_LIST)

    # ─────────────────────────────────────────────
    def extract_order_list(self, files: list[str]) -> pd.DataFrame:
        """
        발주서 Excel 파일 목록을 순서대로 읽어 단일 데이터프레임으로 합친다.

        처리 흐름:
          1. load_workbook() 으로 Excel 파일 열기
          2. 발주번호·입고예정일 등 헤더 정보 읽기
          3. DATA_START_ROW 부터 max_row 까지 반복하며 품목 행 파싱
          4. item_no 가 None 이면 빈 행으로 판단하여 건너뜀

        Args:
            files: 발주서 Excel 파일 절대경로 목록

        Returns:
            발주 항목이 담긴 DataFrame (컬럼: Constants.COL_LIST)

        [최적화 포인트]
          - row_idx 를 별도 변수로 관리해 DataFrame.loc 직접 할당.
            append() 를 반복 호출하는 것보다 메모리 재할당 횟수가 적다.
          - 파일별 예외를 개별 처리하여 일부 오류가 전체를 중단시키지 않는다.
        """
        order_df = pd.DataFrame(columns=Constants.COL_LIST)
        sort_idx = 1   # 전체 품목 순번 (파일에 걸쳐 연속)
        row_idx  = 0   # DataFrame 인덱스

        for file in files:
            try:
                workbook  = load_workbook(file)
                worksheet = workbook.active

                # ── 헤더 정보 읽기 ──
                # 발주번호: 정수 변환 (문자열 방지)
                order_no = int(worksheet.cell(
                    row=Constants.ORDER_NO_ROW,
                    column=Constants.ORDER_NO_COL
                ).value)

                # 입고예정일: datetime 또는 문자열로 저장될 수 있으므로
                # 앞 10자리(YYYY-MM-DD)만 문자열로 슬라이싱
                date_value    = worksheet.cell(
                    row=Constants.DATE_ROW,
                    column=Constants.DATE_COL
                ).value
                delivery_date = str(date_value)[:10] if date_value else ""

                # ── 품목 행 순회 ──
                for row in range(Constants.DATA_START_ROW, worksheet.max_row + 1):
                    item_no = worksheet.cell(row=row, column=Constants.ITEM_NO_COL).value

                    # item_no 가 None 이면 빈 행(또는 데이터 끝) → 건너뜀
                    if item_no is not None:
                        sku_id     = int(worksheet.cell(row=row, column=Constants.SKU_ID_COL).value)
                        sku_name   = str(worksheet.cell(row=row, column=Constants.SKU_NAME_COL).value)
                        order_qty  = int(worksheet.cell(row=row, column=Constants.ORDER_QTY_COL).value)

                        # 바코드는 품목 행 바로 아래(+1) 행에 위치
                        barcode    = str(worksheet.cell(
                            row=row + Constants.BARCODE_ROW_OFFSET,
                            column=Constants.BARCODE_COL
                        ).value)
                        center     = str(worksheet.cell(row=row, column=Constants.CENTER_COL).value)
                        order_price = worksheet.cell(row=row, column=Constants.ORDER_PRICE_COL).value

                        # COL_LIST 순서와 동일하게 리스트 구성
                        item_data = [
                            order_no, sort_idx, int(item_no), '',  # 발주번호, 정렬NO, 발주서NO, 박스NO
                            int(sku_id), sku_name, '',             # SKU ID, SKU 이름, 옵션코드
                            order_qty, '', '', barcode, '',         # 발주수량, 확정수량, 비고, Barcode, 쿠팡옵션코드
                            center, delivery_date, '',             # 물류센터, 입고예정일, 출고예정일
                            order_price, '',                       # 발주공급가, 견적공급가
                        ]
                        order_df.loc[row_idx] = item_data
                        sort_idx += 1
                        row_idx  += 1

            except Exception as e:
                # 개별 파일 오류는 콘솔 출력 후 다음 파일로 계속 진행
                print(f"파일 처리 중 오류 발생 ({file}): {e}")
                continue

        return order_df

    # ─────────────────────────────────────────────
    def merge_with_sku(self, order_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        추출된 발주 데이터프레임에 SKU 마스터 정보를 LEFT JOIN 으로 병합한다.

        Args:
            order_df: extract_order_list() 가 반환한 발주 DataFrame

        Returns:
            (result_df, merged_df) 튜플
              - result_df  : Constants.COL_LIST 순서로 컬럼이 재배치된 최종 DF
              - merged_df  : 병합 직후의 중간 DF (카테고리 요약 계산에 사용)

        Raises:
            FileNotFoundError: SKU 리스트 파일이 존재하지 않을 때

        [최적화 포인트]
          - 필요한 컬럼만 슬라이싱해서 merge() 수행 → 불필요한 컬럼 제거로
            메모리 사용량 절감.
          - reindex() 로 COL_LIST 기준 컬럼 정렬 → 컬럼 순서 보장.
        """
        sku_list_path = os.path.join(self.base_path, Constants.SKU_LIST_FILE)

        if not os.path.exists(sku_list_path):
            raise FileNotFoundError(f"SKU 리스트 파일을 찾을 수 없습니다: {sku_list_path}")

        try:
            # ── 발주 DF에서 필요 컬럼만 선택 (옵션코드·쿠팡옵션코드 제외) ──
            order_cols = [
                "발주번호", "정렬NO", "발주서NO", "박스NO", "SKU ID", "SKU 이름",
                "발주수량", "확정수량", "비고", "SKU Barcode", "물류센터",
                "입고예정일", "출고예정일", "발주공급가",
            ]
            order_subset = order_df.loc[:, order_cols]

            # ── SKU 마스터 읽기: 1행을 헤더로 사용하므로 skiprows=1 ──
            sku_df   = pd.read_excel(sku_list_path, skiprows=1)
            sku_cols = ["SKU ID", "쿠팡옵션코드", "옵션코드", "견적공급가", "상품분류"]
            sku_subset = sku_df.loc[:, sku_cols]

            # ── LEFT JOIN: 발주에 있는 SKU ID 를 기준으로 마스터 정보 추가 ──
            merged_df = order_subset.merge(sku_subset, on="SKU ID", how="left")

            # COL_LIST 순서대로 컬럼 재배치 (없는 컬럼은 NaN 으로 채워짐)
            result_df = merged_df.reindex(columns=Constants.COL_LIST)

            self.current_df = result_df  # 클립보드 복사 기능이 참조
            return result_df, merged_df

        except Exception as e:
            print(f"SKU 병합 중 오류 발생: {e}")
            # 오류 발생 시 원본 order_df 를 그대로 반환해 UI 가 중단되지 않게 함
            return order_df, order_df

    # ─────────────────────────────────────────────
    def get_category_summary(self, merged_df: pd.DataFrame) -> list[int]:
        """
        상품분류별 발주수량 합계를 CATEGORY_ORDER 순서로 반환한다.

        처리 흐름:
          1. groupby("상품분류") → 각 분류별 발주수량 합산
          2. CATEGORY_ORDER 기준 DataFrame 과 LEFT JOIN → 누락 분류는 0 처리
          3. 마지막 행(총발주수량)에 전체 합계 설정
          4. 리스트로 변환하여 반환

        Args:
            merged_df: merge_with_sku() 가 반환한 merged_df

        Returns:
            [여자청바지합계, 사입리스트합계, ..., 총발주수량] 정수 리스트 (9개)
        """
        # ── 카테고리별 합산 ──
        category_summary = merged_df.groupby("상품분류", as_index=False)["발주수량"].agg("sum")

        # 정렬 기준 DataFrame 생성
        category_order_df = pd.DataFrame(
            Constants.CATEGORY_ORDER,
            columns=["상품분류", "순서"]
        )

        # LEFT JOIN 후 순서 컬럼 제거, NaN → 0 처리
        result_df = pd.merge(
            category_order_df,
            category_summary,
            how="left"
        ).drop("순서", axis=1).fillna(0)

        # 마지막 행(인덱스 8 = "총발주수량")에 전체 합계 설정
        result_df.iloc[8, 1] = merged_df["발주수량"].sum()

        # 전치 후 두 번째 행(발주수량 행)을 정수 리스트로 변환
        transposed = result_df.T
        volumes = [int(transposed.iloc[1, x]) for x in range(9)]

        return volumes


# ===================== 로딩 화면 클래스 =====================

class LoadingWindow:
    """
    무거운 작업(파일 읽기, Excel 생성 등) 처리 중
    사용자에게 진행 상태를 알리는 모달 팝업 창.

    - 부모 창의 중앙에 자동으로 위치함.
    - grab_set() 으로 모달(다른 창 조작 불가) 설정.
    - 점(.) 애니메이션으로 진행 중임을 시각적으로 표시.

    사용 예:
        loading = LoadingWindow(self.root, "파일 처리 중...")
        # ... 무거운 작업 ...
        loading.close()
    """

    def __init__(self, parent: tk.Tk, message: str = "처리 중입니다..."):
        self.parent = parent
        self.window = Toplevel(parent)
        self.window.title("처리 중")
        self.window.resizable(False, False)
        self.window.configure(bg=Constants.WINDOW_BG)

        # ── 모달 설정: 이 창이 열린 동안 부모 창 입력 차단 ──
        self.window.transient(parent)  # 부모 창에 종속
        self.window.grab_set()         # 포커스 독점

        # ── 부모 창 중앙에 팝업 배치 ──
        window_width, window_height = 400, 150
        px = parent.winfo_x() + (parent.winfo_width()  // 2) - (window_width  // 2)
        py = parent.winfo_y() + (parent.winfo_height() // 2) - (window_height // 2)
        self.window.geometry(f"{window_width}x{window_height}+{px}+{py}")

        # ── 메시지 레이블 ──
        self.message_label = Label(
            self.window,
            text=message,
            font=(Constants.FONT_NAME, 12, 'normal'),
            bg=Constants.WINDOW_BG,
            fg="#333333"
        )
        self.message_label.pack(pady=30)

        # ── 점 애니메이션 레이블 ──
        self.dots_label = Label(
            self.window,
            text=".",
            font=(Constants.FONT_NAME, 20, 'normal'),
            bg=Constants.WINDOW_BG,
            fg=Constants.BTN_BG_BLUE
        )
        self.dots_label.pack()

        self.dots_count = 0
        self.animate_dots()  # 애니메이션 시작

        self.window.update()  # 즉시 렌더링

    def animate_dots(self):
        """
        0.5초마다 점(.) 1~3개를 순환하는 애니메이션.

        after() 를 재귀 호출하여 이벤트 루프 내에서 비동기로 실행된다.
        창이 파괴된 후 after() 콜백이 남아있으면 TclError 가 발생할 수 있으나,
        close() 에서 window.destroy() 시 pending after 도 함께 취소된다.
        """
        dots = "." * ((self.dots_count % 3) + 1)  # 1, 2, 3 반복
        self.dots_label.config(text=dots)
        self.dots_count += 1
        self.window.after(500, self.animate_dots)  # 500ms 후 재호출

    def update_message(self, message: str):
        """
        로딩 창에 표시되는 안내 문구를 갱신한다.

        무거운 작업의 단계별 진행 상황을 실시간으로 알릴 때 사용.

        Args:
            message: 표시할 새 메시지 문자열
        """
        self.message_label.config(text=message)
        self.window.update()

    def close(self):
        """
        로딩 창을 닫는다.

        작업 완료 후 반드시 호출해야 한다.
        try-except 블록의 finally 또는 except 절에서 호출 권장.
        """
        self.window.destroy()


# ===================== 파일명 변경 클래스 =====================

class FileNameChanger:
    """
    쿠팡 발주서 Excel 파일명을 규칙에 맞게 자동 생성하고 변경하는 유틸.

    파일명 형식: MMDD_물류센터_발주번호.xlsx
    예) 0625_인천센터_30012345678.xlsx

    [설계 방침]
      - 모든 메서드가 @staticmethod 이므로 인스턴스 생성 없이 사용 가능.
      - 순수 함수 형태로, 부수효과를 os.rename() 한 곳에만 격리.
    """

    @staticmethod
    def generate_new_filename(file_path: str) -> str:
        """
        발주서 Excel 에서 날짜·물류센터·발주번호를 읽어 새 파일명을 생성한다.

        Args:
            file_path: 원본 발주서 Excel 파일 절대경로

        Returns:
            "MMDD_물류센터_발주번호.xlsx" 형식의 파일명 문자열

        Raises:
            ValueError: Excel 파싱 중 오류 발생 시
        """
        try:
            workbook  = load_workbook(file_path)
            worksheet = workbook.active

            # 입고예정일 → "MMDD" 문자열 추출 (YYYY-MM-DD 에서 [5:7]+[8:10])
            date_value = str(worksheet.cell(
                row=Constants.DATE_ROW,
                column=Constants.DATE_COL
            ).value)
            date_str = date_value[5:7] + date_value[8:10]  # 예: "0625"

            # 물류센터(화물처) 텍스트
            cargo = str(worksheet.cell(
                row=Constants.DATE_ROW,
                column=Constants.CARGO_COL
            ).value)

            # 발주번호
            order_no = str(worksheet.cell(
                row=Constants.ORDER_NO_ROW,
                column=Constants.ORDER_NO_COL
            ).value)

            return f"{date_str}_{cargo}_{order_no}.xlsx"

        except Exception as e:
            raise ValueError(f"파일명 생성 중 오류: {e}")

    @staticmethod
    def rename_file(source_path: str, dest_path: str):
        """
        파일을 source_path 에서 dest_path 로 이름 변경한다.

        Args:
            source_path: 원본 파일 경로
            dest_path  : 새 파일 경로

        Raises:
            FileExistsError: dest_path 에 동일명 파일이 이미 존재하는 경우
        """
        if os.path.exists(dest_path):
            raise FileExistsError(f"동일한 파일명이 이미 존재합니다: {dest_path}")
        os.rename(source_path, dest_path)


# ===================== GUI 메인 클래스 =====================

class OrderListApp:
    """
    쿠팡 로켓배송 발주 수집·박스라벨 출력 프로그램의 메인 GUI 클래스.

    [화면 구성 - 3-Frame 탭 구조]
      frm1 (발주서 리스트 정리)
        - 발주서 파일 다중 선택 → 데이터 추출 → SKU 병합 → 시트 표시 → 클립보드 복사
      frm2 (발주서 파일명 일괄 변경)
        - 파일 다중 선택 → 저장 폴더 선택 → 규칙 기반 파일명 변경 실행
      frm3 (박스라벨 출력)
        - 발주 리스트 선택 → 출력리스트 확인 → 박스라벨 Excel 생성

    [화면 전환 방식]
      frame.tkraise() 로 동일 컨테이너 내 3개 Frame 을 Z-order 로 전환.
      실제 탭 위젯 대신 헤더의 버튼으로 전환하는 커스텀 탭 구조.

    Attributes:
        root (tk.Tk)              : tkinter 루트 창
        base_path (str)           : 유효한 NAS/로컬 루트 경로
        processor (OrderDataProcessor): 데이터 처리 객체
    """

    def __init__(self, root: tk.Tk, base_path: str):
        self.root      = root
        self.base_path = base_path
        self.processor = OrderDataProcessor(base_path)
        self.setup_window()    # 창 기본 설정
        self.create_widgets()  # 모든 위젯 생성
        self.show_frame(self.frm1)  # 초기 화면: frm1 표시

    # ─────────────────────────────────────────────
    # 창 초기화
    # ─────────────────────────────────────────────

    def setup_window(self):
        """
        루트 창의 제목, 크기, 최소 크기, 리사이즈 정책을 설정한다.

        state('zoomed') : 창을 최대화 상태로 시작 (Windows 전용)
        minsize()       : 리사이즈 시 최소 크기 제한
        """
        self.root.title(Constants.WINDOW_TITLE)
        self.root.geometry(Constants.WINDOW_GEOMETRY)
        self.root.state('zoomed')        # 창 최대화 (Windows)
        self.root.minsize(1520, 850)
        self.root.resizable(True, True)  # 가로·세로 모두 리사이즈 허용

    def create_widgets(self):
        """
        프로그램의 모든 GUI 위젯을 순서대로 생성한다.

        호출 순서가 중요하다. (pack 순서가 레이아웃을 결정)
          1. 상단 헤더(항상 표시)
          2. 컨테이너(3개 프레임이 겹쳐 들어가는 영역)
          3. 각 Frame 내부 위젯
          4. 하단 상태바(항상 표시)
        """
        self.create_header()     # ① 최상단 헤더 영역
        self.create_container()  # ② 3-Frame 컨테이너
        self.create_frame1()     # ③ 발주서 리스트 정리 화면
        self.create_frame2()     # ④ 파일명 변경 화면
        self.create_frame3()     # ⑤ 박스라벨 출력 화면
        self.create_bottom_bar() # ⑥ 최하단 상태 메시지 바

    # ─────────────────────────────────────────────
    # 공통 UI 컴포넌트
    # ─────────────────────────────────────────────

    def create_header(self):
        """
        창 최상단 헤더 프레임을 생성한다.

        - 왼쪽: 프로그램 안내 문구
        - 오른쪽: 화면 전환 버튼 + 종료 버튼

        [버튼 목록]
          발주서 리스트 정리  → frm1 전환 (파란색)
          발주서 파일명 변경  → frm2 전환 (빨간색)
          박스라벨 출력       → frm3 전환 (초록색)
          프로그램 종료       → root.quit() (짙은 회색)
        """
        self.frm_header = Frame(self.root, bg=Constants.SHEET_HEADER_BG, height=50)
        self.frm_header.pack(fill="x")  # 가로 전체 채움

        # 헤더 내부 컨테이너 (좌우 패딩 적용)
        header_content = Frame(self.frm_header, bg=Constants.SHEET_HEADER_BG)
        header_content.pack(fill="both", expand=True, padx=20, pady=10)

        # ── 안내 문구 (왼쪽) ──
        header_text = (
            "★ 프로그램 정보 : 상품정보 수집 오류 시 연동데이터 파일를 확인해 주세요 / "
            "프로그램 연동데이터 : 서버 NAS451/team451/DB/쿠팡SKU리스트.xlsx"
        )
        Label(
            header_content,
            text=header_text,
            bg=Constants.SHEET_HEADER_BG,
            fg="#FFFFFF",
            font=(Constants.FONT_NAME, 9),
            anchor="w"
        ).pack(side="left", fill="x")

        # ── 버튼 그룹 (오른쪽) ──
        btn_frame = Frame(header_content, bg=Constants.SHEET_HEADER_BG)
        btn_frame.pack(side="right")

        # 종료 버튼
        Button(
            btn_frame, text="프로그램 종료", width=15, cursor="hand2",
            relief=Constants.BTN_RELIEF, fg=Constants.BTN_FG, bg=Constants.BTN_BG,
            font=(Constants.FONT_NAME, 9, "bold"),
            command=self.root.quit
        ).pack(side="right", padx=5)

        # 박스라벨 출력 버튼 (frm3)
        Button(
            btn_frame, text="박스라벨 출력", width=16, cursor="hand2",
            relief=Constants.BTN_RELIEF, fg=Constants.BTN_FG, bg=Constants.BTN_BG_GREEN,
            font=(Constants.FONT_NAME, 9, "bold"),
            command=lambda: self.show_frame(self.frm3)
        ).pack(side="right", padx=5)

        # 발주서 리스트 정리 버튼 (frm1)
        Button(
            btn_frame, text="발주서 리스트 정리", width=18, cursor="hand2",
            relief=Constants.BTN_RELIEF, fg=Constants.BTN_FG, bg=Constants.BTN_BG_BLUE,
            font=(Constants.FONT_NAME, 9, "bold"),
            command=lambda: self.show_frame(self.frm1)
        ).pack(side="right", padx=5)

        # 파일명 일괄 변경 버튼 (frm2)
        Button(
            btn_frame, text="발주서 파일명 일괄 변경", width=22, cursor="hand2",
            relief=Constants.BTN_RELIEF, fg=Constants.BTN_FG, bg=Constants.BTN_BG_RED,
            font=(Constants.FONT_NAME, 9, "bold"),
            command=lambda: self.show_frame(self.frm2)
        ).pack(side="right", padx=5)

    def create_container(self):
        """
        3개 Frame(frm1, frm2, frm3)이 겹쳐서 배치되는 컨테이너를 생성한다.

        place(relx=0, rely=0, relwidth=1, relheight=1) 로 각 Frame 이
        컨테이너 전체를 덮도록 배치한 뒤, show_frame() 으로 Z-order 를 바꿔
        원하는 Frame 만 보이도록 한다.
        """
        self.frm_container = Frame(self.root)
        self.frm_container.pack(fill="both", padx=10, pady=5, ipady=5, expand=True)

        # 3개 Frame 생성 (아직 내용은 없음)
        self.frm1 = Frame(self.frm_container)
        self.frm2 = Frame(self.frm_container)
        self.frm3 = Frame(self.frm_container)

        # 모든 Frame 을 컨테이너 전체 크기로 겹쳐 배치
        for frame in (self.frm1, self.frm2, self.frm3):
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)

    def create_bottom_bar(self):
        """
        창 최하단에 상태 메시지를 표시하는 상태 바를 생성한다.

        update_status() 메서드로 메시지를 갱신한다.
        오류 메시지는 빨간색, 일반 메시지는 흰색으로 표시.
        """
        self.frm_bottom = Frame(self.root, bg=Constants.SHEET_HEADER_BG, height=30)
        self.frm_bottom.pack(side="bottom", fill="x")

        self.status_label = Label(
            self.frm_bottom,
            text="준비",
            font=(Constants.FONT_NAME, 10, 'normal'),
            bg=Constants.SHEET_HEADER_BG,
            fg="#FFFFFF",
            anchor="w"
        )
        self.status_label.pack(fill="both", padx=10, pady=5)

    def update_status(self, message: str, is_error: bool = False):
        """
        하단 상태 바의 메시지를 갱신한다.

        Args:
            message : 표시할 메시지 문자열
            is_error: True 이면 빨간색, False 이면 흰색으로 표시
        """
        color = "#FF4000" if is_error else "#FFFFFF"
        self.status_label.config(text=message, fg=color)
        self.root.update()  # 즉시 화면에 반영

    def show_frame(self, frame: Frame):
        """
        지정된 Frame 을 최상위(앞면)로 올려 화면에 표시한다.

        tkraise() 는 해당 Frame 의 Z-order 를 가장 앞으로 바꾼다.
        나머지 Frame 은 뒤로 밀려 보이지 않게 된다.

        Args:
            frame: 표시할 Frame 객체 (self.frm1, frm2, frm3 중 하나)
        """
        frame.tkraise()

    def output_sheet(self, frame: Frame, dataframe: pd.DataFrame):
        """
        tksheet.Sheet 위젯을 생성하여 지정된 Frame 안에 데이터프레임을 표시한다.

        기존 Frame 내 위젯을 모두 제거한 뒤 새 Sheet 를 삽입하므로
        데이터 갱신 시 이 메서드를 재호출하면 된다.

        컬럼 정렬 규칙:
          A~E열  : 가운데 정렬 (발주번호, 정렬NO 등 코드성 데이터)
          H~K열  : 가운데 정렬 (수량, 바코드 등)
          M~O열  : 가운데 정렬 (물류센터, 날짜 등)
          P~Q열  : 오른쪽 정렬 (금액 데이터)

        Args:
            frame    : Sheet 를 배치할 부모 Frame
            dataframe: 표시할 pandas DataFrame
        """
        # 기존 위젯 전부 제거 (Sheet 재생성)
        for widget in frame.winfo_children():
            widget.destroy()

        sheet_data = dataframe.values.tolist()
        col_list   = list(dataframe.columns)

        # ── Sheet 위젯 생성 (frame에 직접 pack — grid/pack 혼용 제거) ──
        sheet_widget = Sheet(
            frame,
            data=sheet_data,
            height=Constants.SHEET_HEIGHT,
            headers=col_list,
            header_height=Constants.SHEET_HEADER_HEIGHT,
            header_fg=Constants.SHEET_HEADER_FG,
            header_bg=Constants.SHEET_HEADER_BG,
        )
        sheet_widget.header_font((Constants.FONT_NAME, Constants.FONT_SIZE_HEADER, 'normal'))
        sheet_widget.font((Constants.FONT_NAME, Constants.FONT_SIZE_NORMAL, 'normal'))
        sheet_widget.table_align(align="w")   # 기본 왼쪽 정렬

        # 특정 컬럼 범위별 정렬 재정의
        sheet_widget["A:E"].align("c")   # 코드성 데이터 → 중앙
        sheet_widget["H:K"].align("c")   # 수량·바코드 → 중앙
        sheet_widget["M:O"].align("c")   # 물류센터·날짜 → 중앙
        sheet_widget["P:Q"].align("e")   # 금액 → 오른쪽

        sheet_widget.set_all_column_widths(
            width=None,
            only_set_if_too_small=False,
            redraw=True,
            recreate_selection_boxes=True,
        )
        sheet_widget.enable_bindings()  # 셀 선택·복사 등 기본 인터랙션 활성화
        sheet_widget.pack(fill="both", expand=True)

    # ─────────────────────────────────────────────
    # Frame1: 발주서 리스트 정리
    # ─────────────────────────────────────────────

    def create_frame1(self):
        """
        발주서 리스트 정리 화면(frm1)의 위젯을 생성한다.

        [레이아웃 구조]
          frm1
          ├── frm1_top (상단 영역)
          │   ├── frm1_left  : 발주서 파일 목록 리스트박스 + 스크롤바
          │   └── frm1_right : 액션 버튼 행 + 카테고리 요약 그리드
          └── frm1_contents  : tksheet 스프레드시트 (데이터 표시)

        [버튼 동작]
          1. 발주서 파일 선택  → frm1_add_file()
          2. 상품정보 수집 실행 → frm1_start()
          3. 수집 상품리스트 복사 → frm1_order_list_copy()
          선택 삭제            → frm1_del_file()
          초기화               → frm1_reset()
        """
        # ── 상단: 파일 목록 + 버튼 영역 ──
        frm1_top = Frame(self.frm1, bg=Constants.WINDOW_BG)
        frm1_top.pack(fill="x", padx=20, pady=15)

        # ── 왼쪽: 파일 목록 리스트박스 ──
        frm1_left = Frame(frm1_top, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm1_left.pack(side="left", fill="both", expand=True, padx=(0, 10))

        frm1_scrollbar = Scrollbar(frm1_left)
        frm1_scrollbar.pack(side="right", fill="y", padx=2, pady=2)

        # 발주서 파일 경로가 표시되는 리스트박스
        # selectmode="extended": Ctrl/Shift 클릭으로 다중 선택 가능
        self.frm1_load_file_list_box = Listbox(
            frm1_left,
            selectmode="extended",
            height=8,
            yscrollcommand=frm1_scrollbar.set,
            relief="flat",
            width=100,
            bd=0,
        )
        self.frm1_load_file_list_box.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        frm1_scrollbar.config(command=self.frm1_load_file_list_box.yview)

        # ── 오른쪽: 버튼 + 카테고리 요약 그리드 ──
        frm1_right = Frame(frm1_top, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm1_right.pack(side="right", fill="both", expand=True, padx=(10, 0))

        # 공통 버튼 스타일 딕셔너리 (반복 코드 제거)
        btn_style = {
            "relief": Constants.BTN_RELIEF,
            "bg":     Constants.BTN_BG,
            "fg":     Constants.BTN_FG,
            "font":   (Constants.FONT_NAME, 9),
            "cursor": "hand2",
        }

        frm1_btns = Frame(frm1_right, bg=Constants.WINDOW_BG)
        frm1_btns.pack(fill="x", padx=10, pady=15)

        # 버튼은 side="right" 로 역순 배치 → 화면에는 좌→우 순서로 표시됨
        Button(frm1_btns, text="초기화",               width=8,  command=self.frm1_reset,            **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="선택 삭제",             width=10, command=self.frm1_del_file,          **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="3. 발주서 수집 상품리스트 복사", width=25, command=self.frm1_order_list_copy, **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="2. 발주서 상품정보 수집 실행",   width=25, command=self.frm1_start,           **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm1_btns, text="1. 발주서 파일 선택",   width=18, command=self.frm1_add_file,          **btn_style).pack(side="right", padx=5, ipady=3)

        # ── 카테고리 요약 그리드 (2행: 헤더 + 값) ──
        frm1_info = Frame(frm1_right, bg="#E0E0E0")  # 배경색이 그리드 라인처럼 보임
        frm1_info.pack(fill="x", padx=10, pady=(0, 15))

        self.info_labels: dict[str, Label] = {}  # 값 레이블 참조 저장
        cols = Constants.INFO_LIST

        # 각 컬럼이 균등하게 공간을 나눠 가짐
        for i in range(len(cols)):
            frm1_info.grid_columnconfigure(i, weight=1)

        # 헤더 행 (row=0)
        for i, col in enumerate(cols):
            Label(
                frm1_info, text=col,
                bg="#EEEEEE", fg="#333333",
                font=(Constants.FONT_NAME, 9, "bold"),
                relief="flat", height=2, width=5,
            ).grid(row=0, column=i, sticky="nswe", padx=1, pady=1)

        # 값 행 (row=1): 초기값 "0", 수집 완료 후 update_status 로 갱신됨
        for i, col in enumerate(cols):
            lbl = Label(
                frm1_info, text="0",
                bg="#FFFFFF", fg="#333333",
                font=(Constants.FONT_NAME, 9),
                relief="flat", height=2, width=5,
            )
            lbl.grid(row=1, column=i, sticky="nswe", padx=1, pady=(0, 1))
            self.info_labels[col] = lbl  # 나중에 값 업데이트를 위해 저장

        # ── 하단: 데이터 시트 영역 ──
        frm1_contents = Frame(self.frm1, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm1_contents.pack(fill="both", padx=20, pady=(0, 20), expand=True)

        self.frm1_xsheet = Frame(frm1_contents)
        self.frm1_xsheet.pack(fill="both", padx=5, pady=5, expand=True)

        # 초기 빈 시트 표시 (컬럼 헤더만 있는 빈 데이터프레임)
        self.output_sheet(self.frm1_xsheet, pd.DataFrame(columns=Constants.COL_LIST))

    def create_frame2(self):
        """
        발주서 파일명 일괄 변경 화면(frm2)의 위젯을 생성한다.

        [레이아웃 구조]
          frm2
          ├── frm2_top     : 안내 문구 + 액션 버튼들
          ├── frm2_path_bar: 저장 폴더 경로 입력창
          └── frm2_contents
              ├── frm2_sector1 : 원본 파일명 리스트박스
              └── frm2_sector2 : 변경된 파일명 리스트박스

        [버튼 동작]
          1. 변경파일 선택    → frm2_add_file()
          2. 저장 폴더 선택   → frm2_output_dir_path()
          3. 파일명 변경 실행 → frm2_change_filename()
          선택 삭제           → frm2_del_file()
          초기화              → frm2_reset()
        """
        # ── 상단 컨트롤 바 ──
        frm2_top = Frame(self.frm2, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_top.pack(fill="x", padx=20, pady=15)

        Label(
            frm2_top,
            text=(
                "★ 파일명 변경 : 쿠팡 로켓배송 발주서 파일을 선택해 주세요. "
                "(변경 파일명 형식 : 입고예정일_물류센터_발주서번호)"
            ),
            font=(Constants.FONT_NAME, 9),
            bg=Constants.WINDOW_BG,
            anchor="w",
        ).pack(side="left", padx=10, pady=10)

        btn_style = {
            "relief": Constants.BTN_RELIEF,
            "bg":     Constants.BTN_BG,
            "fg":     Constants.BTN_FG,
            "font":   (Constants.FONT_NAME, 9),
            "cursor": "hand2",
        }

        Button(frm2_top, text="초기화",           width=8,  command=self.frm2_reset,           **btn_style).pack(side="right", padx=5, pady=8, ipady=3)
        Button(frm2_top, text="선택 삭제",         width=12, command=self.frm2_del_file,         **btn_style).pack(side="right", padx=5, pady=8, ipady=3)
        Button(frm2_top, text="3. 파일명 변경 실행", width=18, command=self.frm2_change_filename, **btn_style).pack(side="right", padx=5, pady=8, ipady=3)
        Button(frm2_top, text="2. 저장 폴더 선택",  width=15, command=self.frm2_output_dir_path, **btn_style).pack(side="right", padx=5, pady=8, ipady=3)
        Button(frm2_top, text="1. 변경파일 선택",   width=15, command=self.frm2_add_file,         **btn_style).pack(side="right", padx=5, pady=8, ipady=3)

        # ── 저장 경로 표시 바 ──
        frm2_path_bar = Frame(self.frm2, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_path_bar.pack(fill="x", padx=20, pady=(0, 15))

        Label(
            frm2_path_bar,
            text="변경파일 저장 폴더",
            font=(Constants.FONT_NAME, 9, "bold"),
            bg=Constants.WINDOW_BG,
        ).pack(side="left", padx=15, pady=10)

        # 저장 폴더 경로를 표시하는 읽기 전용 텍스트 입력창
        self.frm2_txt_save_path = Entry(frm2_path_bar, relief="flat", bg="#F9F9F9")
        self.frm2_txt_save_path.pack(side="left", fill="x", expand=True, padx=(0, 15), pady=10, ipady=3)

        # ── 리스트박스 영역 (원본 | 변경) ──
        frm2_contents = Frame(self.frm2, bg=Constants.WINDOW_BG)
        frm2_contents.pack(fill="both", padx=20, pady=(0, 20), expand=True)

        frm2_sector1 = Frame(frm2_contents, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_sector1.pack(side="left", fill="both", expand=True, padx=(0, 10))

        frm2_sector2 = Frame(frm2_contents, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm2_sector2.pack(side="right", fill="both", expand=True, padx=(10, 0))

        def create_listbox(parent: Frame) -> Listbox:
            """스크롤바가 달린 리스트박스를 생성하는 내부 헬퍼 함수"""
            sb = Scrollbar(parent)
            sb.pack(side="right", fill="y", padx=2, pady=2)
            lb = Listbox(
                parent,
                selectmode="extended",
                yscrollcommand=sb.set,
                relief="flat",
                bd=0,
                highlightthickness=0,
            )
            lb.pack(side="left", fill="both", expand=True, padx=5, pady=5)
            sb.config(command=lb.yview)
            return lb

        self.frm2_original_file = create_listbox(frm2_sector1)  # 원본 파일명 목록
        self.frm2_changed_file  = create_listbox(frm2_sector2)  # 변경 완료 파일명 목록

    def create_frame3(self):
        """
        박스라벨 출력 화면(frm3)의 위젯을 생성한다.

        [레이아웃 구조]
          frm3
          ├── frm3_top      : 안내 문구 + 액션 버튼들
          ├── frm3_contents : tksheet (출력 리스트 미리보기)
          └── frm3_result   : 요약 그리드 (박스수량, 파렛트수량, 출고상품수량)

        [버튼 동작]
          1. 발주리스트 선택  → frm3_load_order()
          2. 출력리스트 확인  → frm3_analyze_order()
          3. 출력파일 생성    → frm3_generate_excel()
          초기화              → frm3_reset()

        [인스턴스 변수]
          frm3_template_path : 박스라벨 템플릿 Excel 경로 (없으면 None)
          frm3_order_path    : 선택된 발주 리스트 파일 경로
          frm3_df            : 분석된 박스 출력 데이터프레임
          frm3_sheet         : tksheet Sheet 위젯 참조
        """
        # ── 상태 변수 초기화 ──
        self.frm3_template_path = get_box_template_path()  # 템플릿 파일 경로
        self.frm3_order_path    = None   # 사용자가 선택한 발주 리스트 경로
        self.frm3_df            = None   # 분석된 데이터프레임
        self.frm3_sheet         = None   # Sheet 위젯 참조

        # ── 상단 안내 + 버튼 ──
        frm3_top = Frame(self.frm3, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm3_top.pack(fill="x", padx=20, pady=15)

        Label(
            frm3_top,
            text=(
                "★ 박스라벨 출력 : 박스 분류가 완료된 발주리스트 파일를 선택 후 출력리스트를 확인하고 "
                "출력파일을 생성합니다. / 출력양식 : " + Constants.BOX_TEMPLATE_REL
            ),
            font=(Constants.FONT_NAME, 9),
            bg=Constants.WINDOW_BG,
            anchor="w",
        ).pack(side="left", padx=10, pady=10)

        frm3_btns = Frame(frm3_top, bg=Constants.WINDOW_BG)
        frm3_btns.pack(side="right", padx=10, pady=8)

        btn_style = {
            "relief": Constants.BTN_RELIEF,
            "bg":     Constants.BTN_BG,
            "fg":     Constants.BTN_FG,
            "font":   (Constants.FONT_NAME, 9),
            "cursor": "hand2",
        }

        Button(frm3_btns, text="초기화",           width=8,  command=self.frm3_reset,           **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm3_btns, text="3. 출력파일 생성",  width=16, command=self.frm3_generate_excel,  **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm3_btns, text="2. 출력리스트 확인", width=16, command=self.frm3_analyze_order,  **btn_style).pack(side="right", padx=5, ipady=3)
        Button(frm3_btns, text="1. 발주리스트 선택", width=16, command=self.frm3_load_order,     **btn_style).pack(side="right", padx=5, ipady=3)

        # ── 시트 영역 ──
        frm3_contents = Frame(self.frm3, bg=Constants.WINDOW_BG, bd=1, relief="solid")
        frm3_contents.pack(fill="both", padx=20, pady=(0, 10), expand=True)

        self.frm3_xsheet = Frame(frm3_contents)
        self.frm3_xsheet.pack(fill="both", padx=5, pady=5, expand=True)
        self._frm3_init_sheet()  # 초기 빈 시트 생성

        # ── 결과 요약 그리드 (하단) ──
        frm3_result = Frame(self.frm3, bg="#E0E0E0", bd=1, relief="solid")
        frm3_result.pack(fill="x", padx=20, pady=(0, 20))

        self.frm3_result_labels: dict[str, Label] = {}
        for i, title in enumerate(Constants.BOX_RESULT_TITLES):
            frm3_result.grid_columnconfigure(i, weight=1)
            Label(
                frm3_result, text=title,
                bg="#EEEEEE", fg="#333333",
                font=(Constants.FONT_NAME, 9, "bold"),
                relief="flat", height=2,
            ).grid(row=0, column=i, sticky="nswe", padx=1, pady=1)

            val_lbl = Label(
                frm3_result, text="-",
                bg="#FFFFFF", fg="#333333",
                font=(Constants.FONT_NAME, 11),
                relief="flat", height=2,
            )
            val_lbl.grid(row=1, column=i, sticky="nswe", padx=1, pady=(0, 1))
            self.frm3_result_labels[title] = val_lbl

        self._frm3_check_template()  # 템플릿 파일 존재 여부 확인 및 경고

    # ── frm3 내부 헬퍼 메서드 ──

    def _frm3_init_sheet(self):
        """
        박스라벨 출력용 빈 tksheet 를 초기화한다.

        기존 위젯을 제거하고 BOX_LABEL_COL_WIDTHS 기준 컬럼으로
        새 Sheet 를 생성한다. 초기화(frm3_reset) 또는 화면 최초 진입 시 호출.
        """
        for widget in self.frm3_xsheet.winfo_children():
            widget.destroy()

        cols = list(Constants.BOX_LABEL_COL_WIDTHS.keys())
        self.frm3_sheet = Sheet(
            self.frm3_xsheet,
            data=[],
            headers=cols,
            height=Constants.SHEET_HEIGHT,
            header_height=Constants.SHEET_HEADER_HEIGHT,
            header_fg=Constants.SHEET_HEADER_FG,
            header_bg=Constants.SHEET_HEADER_BG,
            default_column_width=100,
        )
        self.frm3_sheet.header_font((Constants.FONT_NAME, Constants.FONT_SIZE_HEADER, 'normal'))
        self.frm3_sheet.font((Constants.FONT_NAME, Constants.FONT_SIZE_NORMAL, 'normal'))
        self.frm3_sheet.table_align(align="w")
        self.frm3_sheet.enable_bindings()
        self.frm3_sheet.pack(fill="both", expand=True)

    def _frm3_check_template(self):
        """
        박스라벨 템플릿 Excel 파일의 존재 여부를 확인하고,
        없으면 사용자에게 경고 메시지를 표시한다.

        화면 진입 시(create_frame3)와 초기화 시(frm3_reset) 호출.
        """
        if not self.frm3_template_path:
            expected = " / ".join(
                str(Path(p) / Constants.BOX_TEMPLATE_REL) for p in Constants.DIR_PATHS
            )
            messagebox.showerror(
                "템플릿 오류",
                f"박스라벨 출력양식을 찾을 수 없습니다.\n\n"
                f"아래 경로 중 하나에 파일이 있어야 합니다:\n{expected}",
            )

    def _frm3_apply_column_widths(self, columns: list[str] | None = None):
        """
        frm3_sheet 의 각 컬럼 너비를 BOX_LABEL_COL_WIDTHS 기준으로 설정한다.

        Args:
            columns: 너비를 적용할 컬럼 이름 목록.
                     None 이면 BOX_LABEL_COL_WIDTHS 전체 키를 사용.
        """
        if self.frm3_sheet is None:
            return
        columns = columns or list(Constants.BOX_LABEL_COL_WIDTHS.keys())
        if self.frm3_sheet.total_columns() < len(columns):
            return  # 시트 컬럼 수가 부족하면 스킵
        for idx, col in enumerate(columns):
            width = Constants.BOX_LABEL_COL_WIDTHS.get(col, 100)
            self.frm3_sheet.column_width(column=idx, width=width, redraw=False)
        self.frm3_sheet.redraw()  # 한 번에 렌더링 (개별 redraw 보다 효율적)

    def _frm3_update_result_summary(self):
        """
        frm3 하단 결과 요약 그리드(박스수량/파렛트수량/출고상품수량)를 갱신한다.

        frm3_df 가 없거나 비어있으면 "-" 로 표시한다.
        """
        summary = self._frm3_calc_summary()
        if summary is None:
            values = ("-", "-", "-")
        else:
            last_box, pallets, total_qty = summary
            values = (str(last_box), str(pallets), f"{total_qty:,}")  # 천 단위 콤마

        for title, val in zip(Constants.BOX_RESULT_TITLES, values):
            self.frm3_result_labels[title].config(text=val)

    def _frm3_calc_summary(self) -> tuple[int, int, int] | None:
        """
        frm3_df 에서 박스수량, 파렛트수량, 출고상품수량을 계산한다.

        Returns:
            (last_box, pallets, total_qty) 튜플, 또는 데이터 없으면 None

        [계산 방법]
          last_box  : 박스NO 최댓값 (= 마지막 박스 번호)
          box_count : 고유 박스NO 수 (nunique)
          pallets   : ceil(box_count / BOXES_PER_PALLET)
          total_qty : 수량 컬럼 합계
        """
        if self.frm3_df is None or self.frm3_df.empty:
            return None
        last_box  = int(self.frm3_df['박스NO'].max())
        box_count = self.frm3_df['박스NO'].nunique()
        pallets   = math.ceil(box_count / Constants.BOXES_PER_PALLET)
        total_qty = int(self.frm3_df['수량'].sum())
        return last_box, pallets, total_qty

    def frm3_load_order(self):
        """
        '1. 발주리스트 선택' 버튼 핸들러.

        파일 탐색기로 Excel 파일을 선택하면 경로를 frm3_order_path 에 저장한다.
        시트와 요약 그리드는 초기화되어 이전 분석 결과가 지워진다.
        """
        path = filedialog.askopenfilename(
            title="발주리스트를 선택하세요",
            filetypes=[("Excel", "*.xlsm *.xlsx")],
            initialdir=os.path.join(self.base_path, Constants.ORDER_LIST_DIR) if self.base_path else "",
        )
        if not path:
            return  # 취소 시 아무것도 안 함

        self.frm3_order_path = path
        self.frm3_df = None  # 이전 분석 데이터 초기화

        # 시트 헤더만 유지하고 데이터 비우기
        if self.frm3_sheet:
            self.frm3_sheet.set_sheet_data([])
            self.frm3_sheet.headers(list(Constants.BOX_LABEL_COL_WIDTHS.keys()))

        self._frm3_update_result_summary()
        self.update_status(f"발주리스트 선택: {Path(path).name}")

    def frm3_reset(self):
        """
        '초기화' 버튼 핸들러.

        선택된 발주 리스트와 분석 결과를 모두 초기화한다.
        데이터가 있을 경우 확인 다이얼로그를 표시한다.
        """
        if self.frm3_order_path or self.frm3_df is not None:
            if not messagebox.askyesno("초기화", "발주리스트와 출력리스트를 초기화할까요?"):
                return  # 사용자가 취소한 경우

        self.frm3_order_path    = None
        self.frm3_df            = None
        self.frm3_template_path = get_box_template_path()  # 템플릿 재검색
        self._frm3_init_sheet()                            # 시트 초기화
        self._frm3_update_result_summary()                 # 요약 초기화
        self.update_status("박스라벨 출력 화면이 초기화되었습니다.")

    def frm3_analyze_order(self):
        """
        '2. 출력리스트 확인' 버튼 핸들러.

        선택된 발주 리스트 Excel 의 '발주리스트' 시트를 읽어
        박스NO 기준으로 정제한 뒤 frm3_sheet 에 표시한다.

        처리 흐름:
          1. pd.read_excel() 로 '발주리스트' 시트 읽기 (header=1 → 2번째 행이 헤더)
          2. 필요 컬럼만 선택 후 컬럼명 한글로 변경
          3. 박스NO 가 없는 행(NaN) 제거
          4. 박스NO 를 정수로 변환
          5. Sheet 에 데이터 설정 및 컬럼 너비 적용
        """
        if not self.frm3_order_path:
            messagebox.showwarning("경고", "발주리스트를 선택하세요.")
            return

        try:
            raw = pd.read_excel(self.frm3_order_path, sheet_name="발주리스트", header=1)

            # Excel 헤더에 개행 문자(\n)가 포함되어 있으므로 그대로 사용
            cols = [
                '발주번호', '정렬\nNO', '박스\nNO', 'SKU ID', 'SKU 이름',
                '확정\n수량', '물류센터', '입고예정일',
            ]
            self.frm3_df = raw[cols].copy()

            # 컬럼명을 개행 없는 깔끔한 이름으로 변경
            self.frm3_df.columns = [
                '발주번호', '정렬NO', '박스NO', 'SKU_ID', 'SKU_NAME',
                '수량', '물류센터', '입고예정일',
            ]

            # 박스NO 가 없는 행 제거 (헤더 아래 빈 행 또는 미분류 항목)
            self.frm3_df = self.frm3_df.dropna(subset=['박스NO'])
            self.frm3_df['박스NO'] = self.frm3_df['박스NO'].astype(int)

            # Sheet 업데이트
            self.frm3_sheet.set_sheet_data(self.frm3_df.values.tolist())
            self.frm3_sheet.headers(self.frm3_df.columns.tolist())
            self._frm3_apply_column_widths(self.frm3_df.columns.tolist())
            self._frm3_update_result_summary()
            self.update_status(f"출력리스트 확인 완료: {len(self.frm3_df)}건")

        except Exception as e:
            self.update_status(f"출력리스트 확인 오류: {e}", is_error=True)
            messagebox.showerror("오류", f"발주리스트 분석 중 오류 발생:\n{e}")

    @staticmethod
    def _frm3_sanitize_filename(name: str) -> str:
        """
        파일명에 사용할 수 없는 특수문자를 언더스코어(_)로 대체한다.

        Windows 파일 시스템 금지 문자: \\ / : * ? " < > |

        Args:
            name: 원본 문자열 (예: 물류센터 이름)

        Returns:
            정제된 문자열 (파일명으로 사용 가능)
        """
        return re.sub(r'[\\/:*?"<>|]', '_', str(name).strip())

    def _frm3_build_default_filename(self) -> str:
        """
        박스라벨 출력파일의 기본 저장 파일명을 생성한다.

        형식: YYYYMMDD_박스라벨출력(물류센터).xlsx
        예  : 20250625_박스라벨출력(인천센터).xlsx

        frm3_df 의 첫 번째 행을 기준으로 날짜와 물류센터를 추출한다.
        """
        first    = self.frm3_df.iloc[0]
        date_str = pd.to_datetime(first['입고예정일']).strftime('%Y%m%d')
        center   = self._frm3_sanitize_filename(first['물류센터'])
        return f"{date_str}_박스라벨출력({center}).xlsx"

    @staticmethod
    def _frm3_truncate_sku_name(name) -> str:
        """
        SKU 이름을 BOX_SKU_NAME_MAX_LEN 글자 이내로 잘라 반환한다.

        박스라벨의 SKU 이름 셀 너비 제한으로 인해 너무 긴 이름은 잘라야 한다.

        Args:
            name: SKU 이름 (str 또는 NaN)

        Returns:
            잘린 문자열 (NaN 이면 빈 문자열)
        """
        if pd.isna(name):
            return ''
        text = str(name).strip()
        return text[:Constants.BOX_SKU_NAME_MAX_LEN]  # 초과 시 슬라이싱

    @staticmethod
    def _frm3_copy_style(src, dst):
        """
        openpyxl Cell 의 스타일(폰트, 채움색, 테두리, 정렬, 숫자 형식)을
        src 에서 dst 로 복사한다.

        copy() 를 사용하는 이유: openpyxl 스타일 객체는 참조 공유 시
        하나를 수정하면 다른 셀에도 영향을 줄 수 있으므로 얕은 복사로 독립.

        Args:
            src: 원본 openpyxl Cell 객체
            dst: 대상 openpyxl Cell 객체
        """
        dst.font          = copy(src.font)
        dst.fill          = copy(src.fill)
        dst.border        = copy(src.border)
        dst.alignment     = copy(src.alignment)
        dst.number_format = src.number_format  # 문자열이므로 복사 불필요

    def _frm3_get_template_print_ranges(self, tpl_ws) -> list:
        """
        템플릿 시트에서 인쇄 영역(print_area)을 가져온다.

        인쇄 영역이 설정되어 있지 않으면 기본값으로
        1열~4열, 1행~BOX_BLOCK_ROWS 행 범위를 사용한다.

        Args:
            tpl_ws: openpyxl 템플릿 워크시트

        Returns:
            CellRange 객체 목록
        """
        if tpl_ws._print_area and tpl_ws._print_area.ranges:
            return list(tpl_ws._print_area.sorted())
        # 인쇄 영역 미설정 시 기본 범위 반환
        return [CellRange(min_col=1, min_row=1, max_col=4, max_row=Constants.BOX_BLOCK_ROWS)]

    @staticmethod
    def _frm3_copy_sheet_print_settings(tpl_ws, ws):
        """
        템플릿 시트의 인쇄 설정(여백, 용지 크기, 반복 행/열)을 출력 시트에 복사한다.

        Args:
            tpl_ws: 템플릿 워크시트 (openpyxl)
            ws    : 대상 워크시트 (openpyxl)
        """
        ws.page_margins  = copy(tpl_ws.page_margins)
        ws.page_setup    = copy(tpl_ws.page_setup)
        ws.print_options = copy(tpl_ws.print_options)
        if tpl_ws.print_title_rows:
            ws.print_title_rows = tpl_ws.print_title_rows
        if tpl_ws.print_title_cols:
            ws.print_title_cols = tpl_ws.print_title_cols

    def _frm3_apply_print_areas(self, tpl_ws, ws, block_rows_list: list):
        """
        박스별 실제 블록 행 수를 반영하여 인쇄 영역을 동적으로 설정한다.

        SKU 수에 따라 박스마다 블록 행 수가 다를 수 있으므로,
        고정 stride 대신 누적 오프셋(current_row)으로 각 블록의
        인쇄 범위를 계산한다.

        Args:
            tpl_ws          : 템플릿 워크시트
            ws              : 출력 워크시트
            block_rows_list : 박스별 실제 블록 행 수 목록
                              예) [15, 15, 18, 15]
                              SKU 10개 이하 → 15, 초과 시 15 + 초과분
        """
        base_ranges = self._frm3_get_template_print_ranges(tpl_ws)
        areas       = []
        current_row = 0  # 직전 박스 블록까지의 누적 행 수

        for block_rows in block_rows_list:
            for r in base_ranges:
                areas.append(CellRange(
                    min_col=r.min_col,
                    max_col=r.max_col,
                    min_row=r.min_row + current_row,
                    max_row=r.min_row + current_row + block_rows - 1,  # 실제 블록 끝 행
                ).coord)
            current_row += block_rows  # 다음 블록 시작 기준 갱신

        ws.print_area = areas  # 전체 인쇄 영역 한 번에 설정

    def _frm3_copy_template_block(self, tpl_ws, ws, offset_row: int, block_rows: int):
        """
        템플릿의 1개 박스 블록을 출력 시트의 지정된 오프셋 행부터 복사한다.

        [템플릿 행 구조 - 박스라벨_출력양식.xlsx 기준]
          행1      : 여백 (height=9.95)
          행2      : 박스 적재리스트 제목(A~C 병합) + 박스번호(D열)
          행3      : 입고예정일자(A~B 병합) + 날짜/물류센터(C열) + 빈셀(D열)
          행4      : 업 체 명(A~B 병합) + 업체명(C열) + 빈셀(D열)
          행5      : 발주번호(A~B 병합) + 발주번호값(C열) + 빈셀(D열)
          행6      : 컬럼 헤더 (NO/SKU No./SKU NAME/수량)
          행7~14   : 데이터 슬롯 (8행) ← A열 =IF 수식, B열 SKU_ID, C열 SKU_NAME, D열 수량
          행15     : 하단 여백 (height=9.95)

        [확장 시 처리 순서]
          ① 헤더 행 복사 (행1 ~ 행BOX_HEADER_ROWS)
          ② 기본 데이터 슬롯 복사 (행BOX_HEADER_ROWS+1 ~ 행tpl_rows-BOX_FOOTER_ROWS)
          ③ 초과 데이터 행 삽입 (첫 데이터행 스타일 기준으로 빈 행 추가)
          ④ 하단 여백 행 복사 (항상 블록 마지막 행)

        Args:
            tpl_ws     : 템플릿 워크시트
            ws         : 출력 워크시트
            offset_row : 이 블록의 출력 시작 행 (1-based, 누적 오프셋)
            block_rows : 이 박스 블록의 실제 총 행 수 (≥ BOX_BLOCK_ROWS)
        """
        tpl_rows    = Constants.BOX_BLOCK_ROWS   # 15
        header_rows = Constants.BOX_HEADER_ROWS  # 6 (여백1+제목1+메타3+컬럼헤더1)
        footer_rows = Constants.BOX_FOOTER_ROWS  # 1 (하단 여백)
        num_cols    = 4  # 템플릿 열 수 (A~D)

        # ── ① 헤더 행 복사 (템플릿 1 ~ 6행) ──
        for r in range(1, header_rows + 1):
            dest_r = offset_row + r - 1
            for c in range(1, num_cols + 1):
                s = tpl_ws.cell(r, c)
                d = ws.cell(dest_r, c)
                d.value = s.value
                self._frm3_copy_style(s, d)
            ws.row_dimensions[dest_r].height = tpl_ws.row_dimensions[r].height

        # ── ② 기본 데이터 슬롯 복사 (템플릿 7 ~ 14행) ──
        for r in range(header_rows + 1, tpl_rows - footer_rows + 1):
            dest_r = offset_row + r - 1
            for c in range(1, num_cols + 1):
                s = tpl_ws.cell(r, c)
                d = ws.cell(dest_r, c)
                d.value = None  # 데이터는 이후 frm3_generate_excel에서 채움, 수식 제거
                self._frm3_copy_style(s, d)
            ws.row_dimensions[dest_r].height = tpl_ws.row_dimensions[r].height

        # ── ③ 초과 데이터 행 삽입 (데이터가 기본 슬롯 초과 시에만 실행) ──
        # 초과분 행은 템플릿 첫 데이터 행(BOX_HEADER_ROWS+1 = 7행) 스타일 기준
        extra_count = block_rows - tpl_rows  # 확장된 행 수 (0이면 실행 안 함)
        tpl_data_first_row = header_rows + 1  # 7행
        for i in range(extra_count):
            dest_r = offset_row + tpl_rows - footer_rows + i
            for c in range(1, num_cols + 1):
                d = ws.cell(dest_r, c)
                d.value = None
                self._frm3_copy_style(tpl_ws.cell(tpl_data_first_row, c), d)
            ws.row_dimensions[dest_r].height = tpl_ws.row_dimensions[tpl_data_first_row].height

        # ── ④ 하단 여백 행 복사 (템플릿 15행 → 항상 블록 맨 마지막 행) ──
        for fi in range(footer_rows):
            tpl_footer_r = tpl_rows - footer_rows + 1 + fi   # 템플릿 15행
            dest_r       = offset_row + block_rows - footer_rows + fi
            for c in range(1, num_cols + 1):
                s = tpl_ws.cell(tpl_footer_r, c)
                d = ws.cell(dest_r, c)
                d.value = s.value
                self._frm3_copy_style(s, d)
            ws.row_dimensions[dest_r].height = tpl_ws.row_dimensions[tpl_footer_r].height

        # ── 병합 셀 복사 (헤더 범위의 병합만, 데이터 슬롯 제외) ──
        # 템플릿 병합 셀: A2:C2, A3:B3, A4:B4, A5:B5
        for m in tpl_ws.merged_cells.ranges:
            if m.min_row <= header_rows:
                # 헤더 영역 병합
                try:
                    ws.merge_cells(
                        start_row=m.min_row + offset_row - 1,
                        end_row=m.max_row   + offset_row - 1,
                        start_column=m.min_col,
                        end_column=m.max_col,
                    )
                except Exception:
                    pass  # 이미 병합된 경우 무시

    def frm3_generate_excel(self):
        """
        '3. 출력파일 생성' 버튼 핸들러.

        frm3_df 의 박스별 그룹을 순서대로 처리해 템플릿을 반복 복사하고
        각 박스의 발주번호·입고예정일·물류센터·SKU 정보를 채워 넣는다.

        처리 흐름:
          1. 템플릿 Excel 로드
          2. 출력 Workbook 생성
          3. 컬럼 너비 + 인쇄 설정 복사
          4. 박스별 groupby → 블록 복사 → 셀 값 채우기
          5. 인쇄 영역 설정
          6. 파일 저장 다이얼로그 → 저장

        [최적화 포인트]
          - 로딩 창을 팝업해 사용자에게 진행 중임을 알린다.
          - 저장 다이얼로그 표시 전에 로딩 창을 닫아 UI 블로킹을 방지한다.
          - 예외 발생 시에도 loading.close() 를 반드시 호출한다.
        """
        if self.frm3_df is None:
            messagebox.showwarning("경고", "발주리스트를 먼저 선택·확인하세요.")
            return

        # 템플릿 재검색 (혹시 나중에 파일이 생성된 경우 대비)
        if not self.frm3_template_path:
            self.frm3_template_path = get_box_template_path()
        if not self.frm3_template_path:
            messagebox.showerror("오류", "박스라벨 출력양식 파일을 찾을 수 없습니다.")
            return

        loading = LoadingWindow(self.root, "박스라벨 출력파일을 생성하는 중입니다...")
        try:
            # ── 템플릿 로드 ──
            tpl    = load_workbook(self.frm3_template_path)
            tpl_ws = tpl['sheet']  # 템플릿 시트 이름은 'sheet' 고정

            # ── 출력 Workbook 생성 ──
            out = Workbook()
            ws  = out.active
            ws.title = '박스라벨리스트'

            # 컬럼 너비 복사 (A, B, C, D 열)
            for k, dim in tpl_ws.column_dimensions.items():
                ws.column_dimensions[k].width = dim.width

            self._frm3_copy_sheet_print_settings(tpl_ws, ws)

            # ── 박스별 데이터 채우기 ──
            groups = list(self.frm3_df.groupby('박스NO'))

            # 템플릿 블록 구조 상수 (Constants 와 동기화)
            TPL_ROWS      = Constants.BOX_BLOCK_ROWS    # 기본 블록 행 수: 15
            HEADER_ROWS   = Constants.BOX_HEADER_ROWS   # 헤더 행 수: 6
            FOOTER_ROWS   = Constants.BOX_FOOTER_ROWS   # 하단 여백 행 수: 1
            TPL_DATA_ROWS = TPL_ROWS - HEADER_ROWS - FOOTER_ROWS  # 기본 데이터 슬롯: 8행

            block_rows_list = []  # 박스별 실제 블록 행 수 (인쇄 영역 계산용)
            offset_row = 1        # 현재 박스 블록의 출력 시작 행

            for box_no, group in groups:
                sku_count = len(group)  # 이 박스의 SKU 종류 수

                # SKU 수가 기본 데이터 슬롯(8)을 초과하면 블록 행 수를 동적 확장
                # 예) SKU 11개 → block_rows = 15 + (11 - 8) = 18
                block_rows = TPL_ROWS + max(0, sku_count - TPL_DATA_ROWS)
                block_rows_list.append(block_rows)

                # 1) 템플릿 블록 복사 (헤더 → 데이터슬롯 → 초과행 → 하단여백 순서)
                self._frm3_copy_template_block(tpl_ws, ws, offset_row, block_rows)

                # 2) 박스 헤더 정보 채우기 (첫 번째 행 기준)
                first = group.iloc[0]  # 인덱스 0이 첫 번째 행 (iloc[1]은 오류)

                # 행2 (offset_row + BOX_ROW_OFFSET_TITLE): 박스번호 → D열
                ws.cell(offset_row + Constants.BOX_ROW_OFFSET_TITLE,
                        Constants.BOX_COL_BOX_NO).value = box_no

                # 행3 (offset_row + BOX_ROW_OFFSET_DATE): 날짜/물류센터 → C열
                ws.cell(offset_row + Constants.BOX_ROW_OFFSET_DATE,
                        Constants.BOX_COL_DATE_VALUE).value = (
                    f"{pd.to_datetime(first['입고예정일']).strftime('%m월 %d일')} "
                    f"  /  입고처 ({first['물류센터']})"
                )

                # 행5 (offset_row + BOX_ROW_OFFSET_ORDER): 발주번호 → C열
                # 동일 박스에 발주번호가 2개 이상인 경우 중복 없이 모두 표시
                unique_order_nos = group['발주번호'].drop_duplicates().astype(str).tolist()
                ws.cell(offset_row + Constants.BOX_ROW_OFFSET_ORDER,
                        Constants.BOX_COL_ORDER_VALUE).value = "  /  ".join(unique_order_nos)

                # 3) SKU 데이터 행 채우기 (행7 = offset_row + HEADER_ROWS 부터)
                data_start = offset_row + HEADER_ROWS  # 헤더(6행) 다음 행부터
                for i, (_, row) in enumerate(group.iterrows()):
                    dest_r = data_start + i
                    ws.cell(dest_r, 1).value = i + 1          # 행 내 순번 (1부터)
                    ws.cell(dest_r, 2).value = row['SKU_ID']
                    ws.cell(dest_r, 3).value = self._frm3_truncate_sku_name(row['SKU_NAME'])
                    ws.cell(dest_r, 4).value = row['수량']
                    # 데이터 행 스타일: 템플릿 첫 데이터 행(HEADER_ROWS+1 = 7행) 기준
                    for col in range(1, 5):
                        self._frm3_copy_style(tpl_ws.cell(HEADER_ROWS + 1, col), ws.cell(dest_r, col))

                offset_row += block_rows  # 다음 박스 블록 시작 행 갱신

            # ── 전체 인쇄 영역 설정 (박스별 가변 행 수 반영) ──
            self._frm3_apply_print_areas(tpl_ws, ws, block_rows_list)

            loading.close()  # 저장 다이얼로그 표시 전에 로딩 창 닫기

            # ── 저장 다이얼로그 ──
            save_kwargs = {
                'defaultextension': '.xlsx',
                'filetypes':        [('Excel', '*.xlsx')],
                'initialfile':      self._frm3_build_default_filename(),
            }
            if self.frm3_order_path:
                save_kwargs['initialdir'] = str(Path(self.frm3_order_path).parent)

            save = filedialog.asksaveasfilename(**save_kwargs)
            if save:
                out.save(save)
                self.update_status(f"박스라벨 저장 완료: {Path(save).name}")

                # ── 저장 완료 후 미리보기 여부 확인 ──
                if messagebox.askyesno(
                    '저장 완료',
                    f'저장이 완료되었습니다.\n{Path(save).name}\n\n파일을 열어서 확인하시겠습니까?'
                ):
                    try:
                        os.startfile(save)  # Windows 기본 연결 프로그램(Excel)으로 열기
                    except Exception as open_err:
                        messagebox.showwarning("알림", f"파일을 자동으로 열 수 없습니다.\n직접 열어서 확인해 주세요.\n{open_err}")

        except Exception as e:
            loading.close()
            self.update_status(f"박스라벨 생성 오류: {e}", is_error=True)
            messagebox.showerror("오류", f"박스라벨 출력파일 생성 중 오류 발생:\n{e}")

    # ─────────────────────────────────────────────
    # Frame1 이벤트 핸들러
    # ─────────────────────────────────────────────

    def frm1_add_file(self):
        """
        '1. 발주서 파일 선택' 버튼 핸들러.

        파일 탐색기로 .xlsx 파일을 다중 선택하면
        frm1_load_file_list_box 리스트박스에 경로를 추가한다.
        """
        initial_dir = os.path.join(self.base_path, Constants.ORDER_LIST_DIR) if self.base_path else ""
        file_list = filedialog.askopenfilenames(
            title="발주서를 선택해 주세요.",
            filetypes=(("xlsx 파일", "*.xlsx"), ("모든 파일", "*.*")),
            initialdir=initial_dir,
        )
        for file_path in file_list:
            self.frm1_load_file_list_box.insert(tk.END, file_path)

    def frm1_reset(self):
        """
        '초기화' 버튼 핸들러.

        리스트박스 비우기, 카테고리 요약 초기화, 빈 시트로 복원.
        """
        self.frm1_load_file_list_box.delete(0, tk.END)
        if hasattr(self, 'info_labels'):
            for lbl in self.info_labels.values():
                lbl.config(text="0")
        self.output_sheet(self.frm1_xsheet, pd.DataFrame(columns=Constants.COL_LIST))

    def frm1_del_file(self):
        """
        '선택 삭제' 버튼 핸들러.

        리스트박스에서 현재 선택된 항목을 역순으로 삭제한다.

        [역순 삭제 이유]
          인덱스를 정순으로 삭제하면 앞 항목 삭제 후 뒤 항목의
          인덱스가 밀려 잘못된 항목이 삭제될 수 있다.
          역순으로 처리하면 앞 인덱스가 변하지 않아 안전하다.
        """
        for index in reversed(self.frm1_load_file_list_box.curselection()):
            self.frm1_load_file_list_box.delete(index)

    def frm1_start(self):
        """
        '2. 발주서 상품정보 수집 실행' 버튼 핸들러.

        리스트박스에 파일이 없으면 경고 후 종료,
        있으면 frm1_order_list_ext() 를 호출한다.
        """
        if self.frm1_load_file_list_box.size() == 0:
            messagebox.showwarning("경고", "발주서 파일을 추가하세요.")
            return
        self.frm1_order_list_ext()

    def frm1_order_list_ext(self):
        """
        발주 리스트 추출·병합·시트 표시의 전체 워크플로를 실행한다.

        처리 단계:
          1. LoadingWindow 팝업 표시
          2. OrderDataProcessor.extract_order_list() 호출 → Excel 파싱
          3. OrderDataProcessor.merge_with_sku() 호출 → SKU 마스터 병합
          4. output_sheet() 로 결과 시트 갱신
          5. get_category_summary() 로 카테고리 요약 → info_labels 갱신
          6. LoadingWindow 닫기 + 완료 알림

        예외 발생 시 LoadingWindow 를 닫고 오류 메시지를 표시한다.
        """
        files = list(self.frm1_load_file_list_box.get(0, tk.END))

        loading = LoadingWindow(self.root, "발주서 파일을 읽는 중입니다...")
        try:
            # ── 단계 1: 발주서 Excel 파싱 ──
            loading.update_message("발주서 파일을 읽는 중입니다...")
            self.update_status("발주서 파일을 읽는 중입니다...")
            self.root.update()
            order_df = self.processor.extract_order_list(files)

            # ── 단계 2: SKU 마스터 병합 ──
            loading.update_message("SKU 리스트와 병합 중입니다...")
            self.update_status("SKU 리스트와 병합 중입니다...")
            self.root.update()
            result_df, merged_df = self.processor.merge_with_sku(order_df)

            # ── 단계 3: 시트 출력 ──
            loading.update_message("데이터를 정리하는 중입니다...")
            self.update_status("데이터를 정리하는 중입니다...")
            self.root.update()
            self.output_sheet(self.frm1_xsheet, result_df)

            # ── 단계 4: 카테고리 요약 갱신 ──
            volumes = self.processor.get_category_summary(merged_df)
            if hasattr(self, 'info_labels'):
                for col, vol in zip(Constants.INFO_LIST, volumes):
                    if col in self.info_labels:
                        self.info_labels[col].config(text=str(vol))

            loading.close()
            self.update_status("발주리스트 수집이 완료 되었습니다.")
            messagebox.showinfo("알림", "발주리스트 수집이 완료 되었습니다.")

        except Exception as e:
            loading.close()
            self.update_status(f"오류 발생: {e}", is_error=True)
            messagebox.showerror("오류", f"발주리스트 수집 중 오류 발생: {e}")

    def frm1_order_list_copy(self):
        """
        '3. 발주서 수집 상품리스트 복사' 버튼 핸들러.

        processor.current_df 를 클립보드에 탭 구분 텍스트로 복사한다.
        Excel 에 직접 붙여넣기(Ctrl+V)가 가능하다.

        [조건 검사]
          - 리스트박스가 비어있으면 수집 전으로 판단하고 경고
          - current_df 가 비어있으면 복사할 데이터 없음 경고
        """
        if self.frm1_load_file_list_box.size() == 0:
            messagebox.showwarning("경고", "발주서 상품정보가 수집되지 않았습니다.")
            return

        try:
            if self.processor.current_df.empty:
                messagebox.showwarning("경고", "복사할 데이터가 없습니다.")
                return

            # index=False, header=False: 인덱스·컬럼명 제외하고 데이터만 복사
            self.processor.current_df.to_clipboard(index=False, header=False)
            messagebox.showinfo("알림", "발주리스트 복사가 완료되었습니다.")

        except Exception as e:
            messagebox.showerror("오류", f"복사 중 오류 발생: {e}")

    # ─────────────────────────────────────────────
    # Frame2 이벤트 핸들러
    # ─────────────────────────────────────────────

    def frm2_add_file(self):
        """
        '1. 변경파일 선택' 버튼 핸들러.

        xlsx/xls 파일을 다중 선택해 원본 파일 리스트박스에 추가한다.
        """
        file_list = filedialog.askopenfilenames(
            title="변경 할 파일을 선택하세요",
            filetypes=(("XLSX 파일", "*.xlsx"), ("XLS 파일", "*.xls"), ("모든 파일", "*.*")),
            initialdir="다운로드",  # 기본 초기 디렉터리
        )
        for file_path in file_list:
            self.frm2_original_file.insert(tk.END, file_path)

    def frm2_reset(self):
        """
        '초기화' 버튼 핸들러.

        원본·변경 파일 리스트박스와 저장 경로 입력창을 모두 초기화한다.
        """
        self.frm2_original_file.delete(0, tk.END)
        self.frm2_changed_file.delete(0, tk.END)
        self.frm2_txt_save_path.delete(0, tk.END)

    def frm2_del_file(self):
        """
        '선택 삭제' 버튼 핸들러.

        원본 파일 리스트박스에서 선택된 항목을 역순으로 삭제한다.
        (역순 삭제 이유: frm1_del_file 과 동일)
        """
        for index in reversed(self.frm2_original_file.curselection()):
            self.frm2_original_file.delete(index)

    def frm2_output_dir_path(self):
        """
        '2. 저장 폴더 선택' 버튼 핸들러.

        폴더 탐색기로 저장 폴더를 선택하면
        frm2_txt_save_path 입력창에 경로를 채워 넣는다.
        """
        output_path = filedialog.askdirectory(title="저장 할 폴더를 선택해 주세요")
        if output_path:
            self.frm2_txt_save_path.delete(0, tk.END)
            self.frm2_txt_save_path.insert(0, output_path)

    def frm2_change_filename(self):
        """
        '3. 파일명 변경 실행' 버튼 핸들러.

        원본 파일 목록을 순서대로 처리하여 파일명을 변경한다.

        처리 흐름:
          1. 사전 조건 검사 (파일 목록, 저장 경로 비어있으면 경고)
          2. LoadingWindow 팝업
          3. FileNameChanger.generate_new_filename() → 새 파일명 생성
          4. FileNameChanger.rename_file() → 파일 이동/이름 변경
          5. 변경된 파일명을 frm2_changed_file 리스트박스에 추가
          6. 완료 알림

        [예외 처리]
          - FileExistsError: 동일 파일명 존재 → 경고 메시지 + 중단
          - 기타 예외: 오류 메시지 + 중단
          모든 예외 발생 시 LoadingWindow 를 반드시 닫는다.
        """
        if self.frm2_original_file.size() == 0:
            messagebox.showwarning("경고", "파일명을 변경 할 파일을 추가하세요.")
            return
        if not self.frm2_txt_save_path.get():
            messagebox.showwarning("경고", "저장 경로를 선택하세요")
            return

        files       = list(self.frm2_original_file.get(0, tk.END))
        output_path = self.frm2_txt_save_path.get()

        loading = LoadingWindow(self.root, "파일명을 변경하는 중입니다...")
        try:
            total_files = len(files)
            for processed_count, file_path in enumerate(files, start=1):
                try:
                    msg = f"파일명 변경 중... ({processed_count}/{total_files})"
                    loading.update_message(msg)
                    self.update_status(msg)
                    self.root.update()

                    new_filename = FileNameChanger.generate_new_filename(file_path)
                    dest_path    = os.path.join(output_path, new_filename)
                    FileNameChanger.rename_file(file_path, dest_path)

                    self.frm2_changed_file.insert(tk.END, dest_path)

                except FileExistsError as e:
                    loading.close()
                    messagebox.showwarning("경고", str(e))
                    return
                except Exception as e:
                    loading.close()
                    messagebox.showerror("오류", f"파일 처리 중 오류 발생 ({file_path}): {e}")
                    return

            loading.close()
            self.update_status("파일명 변경이 완료 되었습니다.")
            messagebox.showinfo('알림', "파일명 변경이 완료 되었습니다.")

        except Exception as e:
            loading.close()
            self.update_status(f"파일명 변경 중 오류 발생: {e}", is_error=True)
            messagebox.showerror("오류", f"파일명 변경 중 오류 발생: {e}")


# ===================== 메인 실행 =====================

def main():
    """
    프로그램 진입점.

    처리 흐름:
      1. Constants.DIR_PATHS 에서 유효한 경로 탐색
      2. 경로를 찾지 못하면 빈 문자열로 폴백 (일부 기능 제한)
      3. tk.Tk() 루트 창 생성
      4. OrderListApp 인스턴스 생성 (모든 GUI 초기화)
      5. root.mainloop() 로 이벤트 루프 진입 (프로그램 실행)
    """
    base_path = get_valid_path(Constants.DIR_PATHS)

    if base_path is None:
        print("경고: 유효한 네트워크/로컬 경로를 찾을 수 없습니다. 일부 기능이 제한될 수 있습니다.")
        base_path = ""  # 빈 문자열로 폴백 (파일 탐색기는 동작하나 초기 경로 없음)

    root = tk.Tk()
    app  = OrderListApp(root, base_path)  # GUI 초기화 (app 변수 참조 유지 필요)
    root.mainloop()  # 이벤트 루프 시작 (창을 닫을 때까지 블로킹)


if __name__ == "__main__":
    main()