"""
티셔츠 포장 변환 도구
A파일(발주리스트) → B파일(박스 분할 출고리스트) 변환 프로그램
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import math
import os
import sys
from datetime import datetime
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import tksheet
    TKSHEET_OK = True
except ImportError:
    TKSHEET_OK = False

# ─────────────────────────────────────────────
# 디자인 상수
# ─────────────────────────────────────────────
C = {
    "bg":           "#F8F9FA",
    "bg_panel":     "#FFFFFF",
    "toolbar":      "#2C3E50",
    "accent":       "#3498DB",
    "accent_h":     "#2980B9",
    "text":         "#212529",
    "text2":        "#6C757D",
    "text_light":   "#FFFFFF",
    "success":      "#27AE60",
    "warning":      "#F39C12",
    "error":        "#E74C3C",
    "info":         "#2980B9",
    "border":       "#DEE2E6",
    "grid_hdr":     "#EAF2FB",
    "grid_sel":     "#D6EAF8",
    "row_odd":      "#FFFFFF",
    "row_even":     "#F7FBFF",
    "btn_secondary":"#6C757D",
    "btn_danger":   "#E74C3C",
}

FONT_FAMILY = "Malgun Gothic"
RULES_DIR   = os.path.dirname(os.path.abspath(__file__))

# ─────────────────────────────────────────────
# 변환 로직 (core)
# ─────────────────────────────────────────────
def load_packaging_rules(rules_path: str) -> dict:
    """포장규칙 파일 → {상품코드_사이즈: 입고수량} dict"""
    df = pd.read_excel(rules_path, engine="openpyxl")
    df["key"] = df["상품코드"].astype(str) + "_" + df["사이즈"].astype(str)
    return dict(zip(df["key"], df["입고수량"]))


def load_order_sheet(path: str) -> pd.DataFrame:
    """A파일 발주리스트 시트 로드"""
    df = pd.read_excel(path, sheet_name="발주리스트", header=1, engine="openpyxl")
    df.columns = [
        "발주번호", "정렬NO", "발주서NO", "박스NO",
        "SKU_ID", "SKU이름", "옵션코드",
        "발주수량", "확정수량", "BOX",
        "SKU_Barcode", "쿠팡옵션코드", "물류센터",
        "입고예정일", "출고예정일",
        "발주공급가", "견적공급가",
        "발주번호2", "상품코드", "컬러", "사이즈"
    ]
    return df


def transform(df: pd.DataFrame, rule_dict: dict) -> pd.DataFrame:
    """
    A파일 발주수량을 포장규칙에 따라 박스 단위로 분할.

    알고리즘:
    - 아이템 순서대로 박스에 채움
    - 박스 용량 = 박스를 시작한 SKU의 규칙값
    - 새 SKU가 들어올 때 규칙이 더 작으면 박스 용량을 줄임
      (이미 채운 양이 새 용량 초과 → 새 박스로)
    - 수량 > 잔여: 잔여만큼만 현재박스, 나머지는 새 박스
    """
    result_rows = []
    box_num = 1
    box_capacity: Optional[int] = None
    box_used = 0

    for _, row in df.iterrows():
        qty_raw = row["발주수량"]
        if pd.isna(qty_raw):
            continue
        qty = int(qty_raw)
        if qty <= 0:
            continue

        code = str(row["상품코드"])
        size_raw = row["사이즈"]
        size = str(int(size_raw)) if pd.notna(size_raw) else ""
        key = f"{code}_{size}"
        sku_cap = int(rule_dict.get(key, 60))

        remaining = qty
        first_chunk = True

        while remaining > 0:
            # 새 박스 시작
            if box_capacity is None:
                box_capacity = sku_cap
                box_used = 0

            # 박스 용량 동적 조정 (더 작은 규칙 SKU 등장 시)
            if sku_cap < box_capacity:
                if box_used >= sku_cap:
                    box_num += 1
                    box_capacity = sku_cap
                    box_used = 0
                else:
                    box_capacity = sku_cap

            available = box_capacity - box_used
            take = min(remaining, available)

            new_row = row.copy()
            new_row["BOX"] = box_num
            new_row["확정수량"] = take
            new_row["발주수량"] = qty if first_chunk else None
            result_rows.append(dict(new_row))

            box_used += take
            remaining -= take
            first_chunk = False

            if box_used >= box_capacity:
                box_num += 1
                box_capacity = None
                box_used = 0

    result = pd.DataFrame(result_rows)
    # 박스NO 컬럼을 BOX 값으로 채움 (정수)
    result["박스NO"] = result["BOX"].astype(int)
    return result


def export_to_excel(result_df: pd.DataFrame, save_path: str,
                    original_path: str) -> None:
    """변환 결과를 B파일 포맷과 동일하게 저장"""
    # 원본 워크북 로드 (다른 시트 보존)
    try:
        wb = openpyxl.load_workbook(original_path, keep_vba=True)
    except Exception:
        wb = openpyxl.Workbook()

    # 발주리스트 시트 교체
    if "발주리스트" in wb.sheetnames:
        del wb["발주리스트"]
    ws = wb.create_sheet("발주리스트", 0)

    # ── 1행: 타이틀
    total_qty = int(result_df["확정수량"].sum())
    ws["A1"] = "발주서 리스트"
    ws["F1"] = "※ 재고수정 후 출고리스트 생성 버튼클릭해 주세요"
    ws["G1"] = "합계수량 : "
    ws["H1"] = total_qty
    ws["I1"] = total_qty

    # ── 2행: 헤더
    headers = [
        "발주번호", "정렬  NO", "발주서NO", "박스  NO",
        "SKU ID", "SKU 이름", "옵션코드",
        "발주  수량", "확정  수량", "BOX",
        "SKU Barcode", "쿠팡옵션코드", "물류센터",
        "입고예정일", "출고예정일",
        "발주공급가", "견적공급가",
        "발주번호", "상품코드", "컬러", "사이즈"
    ]
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF", name=FONT_FAMILY, size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    # ── 데이터 행 매핑
    col_map = [
        "발주번호", "정렬NO", "발주서NO", "박스NO",
        "SKU_ID", "SKU이름", "옵션코드",
        "발주수량", "확정수량", "BOX",
        "SKU_Barcode", "쿠팡옵션코드", "물류센터",
        "입고예정일", "출고예정일",
        "발주공급가", "견적공급가",
        "발주번호2", "상품코드", "컬러", "사이즈"
    ]

    fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    fill_even = PatternFill("solid", fgColor="F7FBFF")
    data_font = Font(name=FONT_FAMILY, size=9)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, (_, row) in enumerate(result_df.iterrows(), start=3):
        fill = fill_odd if r_idx % 2 == 1 else fill_even
        for c_idx, col_key in enumerate(col_map, start=1):
            val = row.get(col_key)
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            # 날짜 처리
            if col_key in ("입고예정일", "출고예정일") and val is not None:
                try:
                    val = pd.to_datetime(val).date()
                except Exception:
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.border = border
            # 정렬
            if c_idx in (6, 7, 12):  # SKU이름, 옵션코드, 쿠팡옵션코드
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # 열 너비 설정
    col_widths = [14,8,9,8, 12,50,28, 9,9,6, 18,28,8, 12,12, 10,10, 14,10,8,6]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 행 높이
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 30
    for r in range(3, 3 + len(result_df)):
        ws.row_dimensions[r].height = 16

    # 틀 고정 (헤더 2행 고정)
    ws.freeze_panes = "A3"

    try:
        ext = os.path.splitext(save_path)[1].lower()
        if ext == ".xlsm":
            wb.save(save_path)
        else:
            wb.save(save_path)
    except Exception as e:
        # xlsm 저장 실패 시 xlsx로 재시도
        alt_path = save_path.replace(".xlsm", ".xlsx").replace(".xls", ".xlsx")
        if alt_path != save_path:
            wb.save(alt_path)
            raise RuntimeError(f"xlsm 저장 실패, {alt_path}에 저장됨: {e}")
        raise


def make_box_summary(df: pd.DataFrame) -> pd.DataFrame:
    """변환 결과 → 박스별 합계수량 (박스번호 중복 없음)"""
    agg = df.groupby("BOX", as_index=False)["확정수량"].sum()
    agg["BOX"] = agg["BOX"].astype(int)
    agg["확정수량"] = agg["확정수량"].astype(int)
    return agg.sort_values("BOX").reset_index(drop=True)


# ─────────────────────────────────────────────
# tksheet 기반 그리드 또는 fallback Treeview
# ─────────────────────────────────────────────
class DataGrid(tk.Frame):
    """tksheet 사용 가능하면 Sheet, 아니면 Treeview fallback"""

    DISPLAY_COLS = [
        "발주번호", "정렬NO", "발주서NO", "박스NO",
        "SKU이름", "옵션코드",
        "발주수량", "확정수량", "BOX",
        "SKU_Barcode", "상품코드", "컬러", "사이즈"
    ]

    def __init__(self, parent, **kw):
        super().__init__(parent, bg=C["bg_panel"], **kw)
        self._df: Optional[pd.DataFrame] = None
        self._use_sheet = TKSHEET_OK
        self._build()

    def _build(self):
        if self._use_sheet:
            self._sheet = tksheet.Sheet(
                self,
                theme="light blue",
                header_font=(FONT_FAMILY, 9, "bold"),
                font=(FONT_FAMILY, 9, "normal"),
                row_index=True,
                show_x_scrollbar=True,
                show_y_scrollbar=True,
            )
            self._sheet.enable_bindings(
                "single_select", "column_select", "row_select",
                "column_width_resize", "copy", "rc_select"
            )
            self._sheet.pack(fill="both", expand=True)
        else:
            # Treeview fallback
            frame = tk.Frame(self, bg=C["bg_panel"])
            frame.pack(fill="both", expand=True)

            vsb = ttk.Scrollbar(frame, orient="vertical")
            hsb = ttk.Scrollbar(frame, orient="horizontal")
            self._tree = ttk.Treeview(
                frame, show="headings",
                yscrollcommand=vsb.set,
                xscrollcommand=hsb.set,
                selectmode="extended"
            )
            vsb.config(command=self._tree.yview)
            hsb.config(command=self._tree.xview)
            vsb.pack(side="right", fill="y")
            hsb.pack(side="bottom", fill="x")
            self._tree.pack(fill="both", expand=True)

    def load(self, df: pd.DataFrame):
        self._df = df
        cols = [c for c in self.DISPLAY_COLS if c in df.columns]
        display = df[cols].copy()

        # 숫자 정수화
        for c in ["발주수량", "확정수량", "BOX", "박스NO", "정렬NO", "발주서NO"]:
            if c in display.columns:
                display[c] = display[c].apply(
                    lambda x: int(x) if pd.notna(x) else ""
                )

        if self._use_sheet:
            self._sheet.headers(cols)
            data = [list(row) for row in display.itertuples(index=False)]
            self._sheet.set_sheet_data(data)
            self._sheet.refresh()
            # 헤더 색상
            self._sheet.highlight_cells(
                row=0, column=None,
                bg=C["grid_hdr"], fg=C["text"],
                canvas="column_headers"
            )
        else:
            self._tree["columns"] = cols
            for c in cols:
                w = 120 if c in ("SKU이름", "옵션코드") else 70
                self._tree.heading(c, text=c)
                self._tree.column(c, width=w, minwidth=40)
            self._tree.delete(*self._tree.get_children())
            for _, row in display.iterrows():
                vals = [row[c] for c in cols]
                self._tree.insert("", "end", values=vals)

    def row_count(self) -> int:
        return len(self._df) if self._df is not None else 0


class BoxSummaryGrid(DataGrid):
    """박스별 수량 요약 그리드"""
    DISPLAY_COLS = ["BOX", "확정수량"]


# ─────────────────────────────────────────────
# 메인 앱
# ─────────────────────────────────────────────
class App:
    RULES_FILENAME = "티셔츠포장규칙.xlsx"

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("티셔츠 포장 변환 도구")
        self.root.geometry("1600x900+10+10")
        self.root.configure(bg=C["bg"])
        self.root.minsize(1400, 800)
        try:
            self.root.state("zoomed")
        except Exception:
            pass

        self._a_path: Optional[str] = None
        self._rules_path: Optional[str] = None
        self._rules: dict = {}
        self._df_a: Optional[pd.DataFrame] = None
        self._df_result: Optional[pd.DataFrame] = None
        self._tab_mode = "A"   # "A" or "RESULT"

        self._apply_ttk_styles()
        self._build_ui()
        self._try_load_default_rules()

    # ── TTK 스타일 ──────────────────────────────
    def _apply_ttk_styles(self):
        s = ttk.Style()
        s.theme_use("clam")

        s.configure("Toolbar.TFrame",    background=C["toolbar"])
        s.configure("BG.TFrame",         background=C["bg"])
        s.configure("Panel.TFrame",      background=C["bg_panel"])
        s.configure("Status.TFrame",     background=C["bg_panel"],
                    relief="sunken", borderwidth=1)

        s.configure("Primary.TButton",
                    background=C["accent"], foreground=C["text_light"],
                    font=(FONT_FAMILY, 9, "bold"),
                    relief="flat", padding=(12, 6), borderwidth=0)
        s.map("Primary.TButton",
              background=[("active", C["accent_h"]), ("disabled", "#A9CCE3")])

        s.configure("Secondary.TButton",
                    background=C["btn_secondary"], foreground=C["text_light"],
                    font=(FONT_FAMILY, 9),
                    relief="flat", padding=(10, 5))
        s.map("Secondary.TButton",
              background=[("active", "#5A6268"), ("disabled", "#CCC")])

        s.configure("Danger.TButton",
                    background=C["btn_danger"], foreground=C["text_light"],
                    font=(FONT_FAMILY, 9),
                    relief="flat", padding=(10, 5))
        s.map("Danger.TButton",
              background=[("active", "#C0392B")])

        s.configure("Status.TLabel",
                    background=C["bg_panel"],
                    foreground=C["text2"],
                    font=(FONT_FAMILY, 8))

        s.configure("Info.TLabel",
                    background=C["bg"],
                    foreground=C["text"],
                    font=(FONT_FAMILY, 9))

        s.configure("TNotebook",           background=C["bg"])
        s.configure("TNotebook.Tab",
                    background=C["border"], foreground=C["text"],
                    font=(FONT_FAMILY, 9), padding=(12, 5))
        s.map("TNotebook.Tab",
              background=[("selected", C["accent"])],
              foreground=[("selected", C["text_light"])])

        s.configure("Horizontal.TProgressbar",
                    troughcolor=C["border"],
                    background=C["accent"],
                    thickness=12)

    # ── UI 빌드 ─────────────────────────────────
    def _build_ui(self):
        # Zone A: Toolbar
        self.toolbar = ttk.Frame(self.root, style="Toolbar.TFrame", height=80)
        self.toolbar.pack(fill="x", side="top")
        self.toolbar.pack_propagate(False)
        self._build_toolbar()

        # Zone B: 데이터 뷰
        self.view_frame = ttk.Frame(self.root, style="BG.TFrame")
        self.view_frame.pack(fill="both", expand=True, padx=6, pady=(4, 0))
        self._build_view()

        # Zone C: 상태바
        self.status_frame = ttk.Frame(self.root, style="Status.TFrame", height=52)
        self.status_frame.pack(fill="x", side="bottom")
        self.status_frame.pack_propagate(False)
        self._build_statusbar()

    def _build_toolbar(self):
        # 왼쪽 패딩
        pad = tk.Frame(self.toolbar, bg=C["toolbar"], width=12)
        pad.pack(side="left")

        def sep():
            tk.Frame(self.toolbar, bg="#4A6580", width=1).pack(
                side="left", fill="y", padx=8, pady=12)

        # ── 파일 그룹
        self._btn_open = self._tb_btn("📂  A파일 열기", self._open_a_file, "Primary")
        sep()

        # ── 포장규칙
        self._btn_rules = self._tb_btn("📋  포장규칙", self._open_rules, "Secondary")
        self._lbl_rules = tk.Label(
            self.toolbar, text="규칙: 미설정", bg=C["toolbar"],
            fg="#95A5A6", font=(FONT_FAMILY, 8))
        self._lbl_rules.pack(side="left", padx=(0, 4))
        sep()

        # ── 변환
        self._btn_transform = self._tb_btn(
            "🔄  변환 실행", self._run_transform, "Primary")
        self._btn_transform.state(["disabled"])
        sep()

        # ── 저장
        self._btn_save = self._tb_btn("💾  파일 저장", self._save_result, "Secondary")
        self._btn_save.state(["disabled"])
        sep()

        # ── 전체 복사
        self._btn_copy = self._tb_btn("📋  전체복사", self._copy_result, "Secondary")
        self._btn_copy.state(["disabled"])
        sep()

        # ── 초기화
        self._btn_reset = self._tb_btn("🗑  초기화", self._reset, "Danger")

        # 오른쪽: 파일 정보
        self._lbl_file = tk.Label(
            self.toolbar, text="파일: 없음", bg=C["toolbar"],
            fg="#BDC3C7", font=(FONT_FAMILY, 8), anchor="e")
        self._lbl_file.pack(side="right", padx=16)

    def _tb_btn(self, text: str, cmd, style="Primary") -> ttk.Button:
        b = ttk.Button(self.toolbar, text=text, command=cmd,
                       style=f"{style}.TButton")
        b.pack(side="left", padx=(0, 4), pady=16)
        return b

    def _build_view(self):
        # 탭 노트북
        self.notebook = ttk.Notebook(self.view_frame)
        self.notebook.pack(fill="both", expand=True)

        # 탭1: A파일 원본
        tab_a = ttk.Frame(self.notebook, style="Panel.TFrame")
        self.notebook.add(tab_a, text="  📋 A파일 원본  ")
        self._grid_a = DataGrid(tab_a)
        self._grid_a.pack(fill="both", expand=True, padx=2, pady=2)

        # 탭2: 변환 결과
        tab_b = ttk.Frame(self.notebook, style="Panel.TFrame")
        self.notebook.add(tab_b, text="  📦 변환 결과 (B파일)  ")
        self._grid_b = DataGrid(tab_b)
        self._grid_b.pack(fill="both", expand=True, padx=2, pady=2)

        # 탭3: 박스별 수량 요약
        tab_s = ttk.Frame(self.notebook, style="Panel.TFrame")
        self.notebook.add(tab_s, text="  📊 박스별 수량  ")
        self._grid_summary = BoxSummaryGrid(tab_s)
        self._grid_summary.pack(fill="both", expand=True, padx=2, pady=2)

        # 탭 전환 시 info 업데이트
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

    def _build_statusbar(self):
        # 진행바
        self._progress = ttk.Progressbar(
            self.status_frame,
            style="Horizontal.TProgressbar",
            length=220, mode="determinate"
        )
        self._progress.pack(side="left", padx=(10, 6), pady=14)

        # 상태 메시지
        self._lbl_status = ttk.Label(
            self.status_frame, text="준비", style="Status.TLabel")
        self._lbl_status.pack(side="left", padx=4)

        # 오른쪽: 행수/열수 정보
        self._lbl_info = ttk.Label(
            self.status_frame, text="", style="Status.TLabel")
        self._lbl_info.pack(side="right", padx=12)

        # 처리시간
        self._lbl_time = ttk.Label(
            self.status_frame, text="", style="Status.TLabel")
        self._lbl_time.pack(side="right", padx=8)

    # ── 기본 포장규칙 로드 ───────────────────────
    def _try_load_default_rules(self):
        """스크립트 위치 우선, 없으면 실행 디렉토리에서 포장규칙 자동 로드"""
        candidates = [
            os.path.join(RULES_DIR, self.RULES_FILENAME),
            os.path.join(os.path.dirname(sys.argv[0]), self.RULES_FILENAME),
            os.path.join(os.getcwd(), self.RULES_FILENAME),
        ]
        for p in candidates:
            if os.path.exists(p):
                self._load_rules(p)
                break

    def _load_rules(self, path: str):
        try:
            self._rules = load_packaging_rules(path)
            self._rules_path = path
            name = os.path.basename(path)
            self._lbl_rules.config(
                text=f"규칙: {name} ({len(self._rules)}건)",
                fg="#2ECC71")
            self._update_transform_btn()
        except Exception as e:
            self._set_status(f"규칙 파일 오류: {e}", "error")

    # ── 이벤트 핸들러 ────────────────────────────
    def _open_a_file(self):
        path = filedialog.askopenfilename(
            title="A파일 (발주리스트) 선택",
            filetypes=[
                ("Excel 파일", "*.xlsm *.xlsx *.xls"),
                ("모든 파일", "*.*")
            ]
        )
        if not path:
            return
        self._load_a_file(path)

    def _load_a_file(self, path: str):
        def worker():
            t0 = datetime.now()
            self._set_status("⏳ 파일 로딩 중...", "info")
            self._set_progress(30)
            try:
                df = load_order_sheet(path)
                self._set_progress(80)
                self.root.after(0, lambda: self._on_a_loaded(df, path, t0))
            except Exception as e:
                self.root.after(0, lambda: self._set_status(
                    f"❌ 파일 로드 실패: {e}", "error"))
                self.root.after(0, lambda: self._set_progress(0))

        threading.Thread(target=worker, daemon=True).start()

    def _on_a_loaded(self, df: pd.DataFrame, path: str, t0):
        self._df_a = df
        self._a_path = path
        elapsed = (datetime.now() - t0).total_seconds()

        self._grid_a.load(df)
        self.notebook.select(0)

        name = os.path.basename(path)
        self._lbl_file.config(text=f"파일: {name}")
        total_qty = int(df["발주수량"].sum()) if "발주수량" in df.columns else 0
        self._set_status(
            f"✅ 로드 완료 — {len(df):,}행 / 발주수량 합계: {total_qty:,}개",
            "success")
        self._set_progress(100)
        self._lbl_info.config(
            text=f"행: {len(df):,}  |  열: {len(df.columns)}")
        self._lbl_time.config(text=f"처리시간: {elapsed:.2f}s")

        self._update_transform_btn()
        self.root.after(3000, lambda: self._set_progress(0))

    def _open_rules(self):
        path = filedialog.askopenfilename(
            title="포장규칙 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")]
        )
        if path:
            self._load_rules(path)

    def _run_transform(self):
        if self._df_a is None:
            messagebox.showwarning("알림", "A파일을 먼저 불러오세요.")
            return
        if not self._rules:
            messagebox.showwarning("알림", "포장규칙 파일을 먼저 불러오세요.")
            return

        def worker():
            t0 = datetime.now()
            self._set_status("⏳ 변환 중...", "info")
            self._set_progress(20)
            try:
                df_result = transform(self._df_a, self._rules)
                self._set_progress(80)
                self.root.after(0, lambda: self._on_transform_done(df_result, t0))
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                self.root.after(0, lambda: self._set_status(
                    f"❌ 변환 실패: {e}", "error"))
                self.root.after(0, lambda: self._set_progress(0))
                print(tb)

        self._btn_transform.state(["disabled"])
        threading.Thread(target=worker, daemon=True).start()

    def _on_transform_done(self, df_result: pd.DataFrame, t0):
        self._df_result = df_result
        elapsed = (datetime.now() - t0).total_seconds()

        self._grid_b.load(df_result)
        self._grid_summary.load(make_box_summary(df_result))
        self.notebook.select(1)

        total_boxes = int(df_result["BOX"].max()) if "BOX" in df_result.columns else 0
        total_qty   = int(df_result["확정수량"].sum())
        orig_qty    = int(self._df_a["발주수량"].sum())

        self._set_status(
            f"✅ 변환 완료 — {len(df_result):,}행 / {total_boxes:,}박스 / "
            f"확정수량: {total_qty:,}개 (원본: {orig_qty:,}개)",
            "success")
        self._set_progress(100)
        self._lbl_info.config(
            text=f"행: {len(df_result):,}  |  박스: {total_boxes:,}개")
        self._lbl_time.config(text=f"처리시간: {elapsed:.3f}s")

        self._btn_save.state(["!disabled"])
        self._btn_copy.state(["!disabled"])
        self._btn_transform.state(["!disabled"])
        self.root.after(5000, lambda: self._set_progress(0))

    def _save_result(self):
        if self._df_result is None:
            messagebox.showwarning("알림", "먼저 변환을 실행하세요.")
            return

        # 기본 저장 경로
        default_name = "B_" + os.path.basename(self._a_path or "output.xlsx")
        save_path = filedialog.asksaveasfilename(
            title="변환 결과 저장",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[
                ("Excel 매크로 파일", "*.xlsm"),
                ("Excel 파일", "*.xlsx"),
                ("모든 파일", "*.*")
            ]
        )
        if not save_path:
            return

        def worker():
            t0 = datetime.now()
            self._set_status("⏳ 저장 중...", "info")
            self._set_progress(40)
            try:
                export_to_excel(
                    self._df_result,
                    save_path,
                    self._a_path or save_path
                )
                self._set_progress(100)
                elapsed = (datetime.now() - t0).total_seconds()
                self.root.after(0, lambda: self._set_status(
                    f"✅ 저장 완료: {os.path.basename(save_path)}", "success"))
                self.root.after(0, lambda: self._lbl_time.config(
                    text=f"저장: {elapsed:.2f}s"))
            except Exception as e:
                self.root.after(0, lambda: self._set_status(
                    f"❌ 저장 실패: {e}", "error"))
            finally:
                self.root.after(0, lambda: self.root.after(
                    4000, lambda: self._set_progress(0)))

        threading.Thread(target=worker, daemon=True).start()

    def _reset(self):
        if not messagebox.askyesno("초기화", "모든 데이터를 초기화하시겠습니까?"):
            return
        self._df_a = None
        self._df_result = None
        self._a_path = None

        # 그리드 초기화
        for grid in (self._grid_a, self._grid_b, self._grid_summary):
            grid.load(pd.DataFrame())

        self._lbl_file.config(text="파일: 없음")
        self._btn_transform.state(["disabled"])
        self._btn_save.state(["disabled"])
        self._btn_copy.state(["disabled"])
        self._set_status("초기화 완료", "info")
        self._set_progress(0)
        self._lbl_info.config(text="")
        self._lbl_time.config(text="")

    def _on_tab_change(self, event):
        tab = self.notebook.index(self.notebook.select())
        if tab == 0 and self._df_a is not None:
            self._lbl_info.config(
                text=f"행: {len(self._df_a):,}  |  열: {len(self._df_a.columns)}")
        elif tab == 1 and self._df_result is not None:
            boxes = int(self._df_result["BOX"].max())
            self._lbl_info.config(
                text=f"행: {len(self._df_result):,}  |  박스: {boxes:,}개")
        elif tab == 2 and self._df_result is not None:
            summary = make_box_summary(self._df_result)
            boxes = summary["BOX"].nunique()
            self._lbl_info.config(
                text=f"박스: {boxes:,}개  |  SKU행: {len(summary):,}")

    def _copy_result(self):
        if self._df_result is None:
            return
        cols = [c for c in DataGrid.DISPLAY_COLS if c in self._df_result.columns]
        display = self._df_result[cols].copy()
        for c in ["발주수량", "확정수량", "BOX", "박스NO", "정렬NO", "발주서NO"]:
            if c in display.columns:
                display[c] = display[c].apply(lambda x: int(x) if pd.notna(x) else "")
        tsv = "\t".join(cols) + "\n"
        tsv += "\n".join(
            "\t".join(str(v) for v in row)
            for row in display.itertuples(index=False)
        )
        self.root.clipboard_clear()
        self.root.clipboard_append(tsv)
        self._set_status(f"✅ {len(display):,}행 클립보드에 복사됨 (엑셀에 붙여넣기 가능)", "success")

    # ── 유틸 ────────────────────────────────────
    def _set_status(self, msg: str, level: str = "info"):
        colors = {
            "info":    C["info"],
            "success": C["success"],
            "warning": C["warning"],
            "error":   C["error"],
        }
        self._lbl_status.config(text=msg, foreground=colors.get(level, C["text"]))

    def _set_progress(self, value: int):
        self._progress["value"] = max(0, min(100, value))

    def _update_transform_btn(self):
        if self._df_a is not None and self._rules:
            self._btn_transform.state(["!disabled"])
        else:
            self._btn_transform.state(["disabled"])


# ─────────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────────
def main():
    root = tk.Tk()

    # 아이콘 (있으면)
    try:
        root.iconbitmap(default="icon.ico")
    except Exception:
        pass

    # DPI 보정 (Windows)
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
